# -*- coding: utf-8 -*-
"""
拍卖会系统 · Flask 应用
- 本文件在保持原有功能不变的前提下，按“相关性”重排了路由与工具函数的顺序
- 每个 @app.route 上方都注明：
    1) 是否渲染 HTML（对应模板路径）
    2) 功能说明（数据来源 / 读写行为 / 参数要点）
- 仅新增了注释与排版（空行），业务逻辑未改
"""

from flask import Flask, render_template, jsonify, request, abort
from config import DEBUG, HOST, PORT

# 复用你已定义的 SQLAlchemy 模型与会话（在 create_database.py 中）
from create_database import get_session, Item, Seller, Auction, Buyer, OperationLog, MaterialOption, OutboundLog

from decimal import Decimal, InvalidOperation
import os
from werkzeug.utils import secure_filename
import shutil
import json
# —— 导出所需 & 下载 ——
from flask import send_file
from pathlib import Path
import tempfile
import pythoncom
import win32com.client as win32

import io
from datetime import datetime

# 可选依赖：没装也能运行，但导出会报“缺依赖”的错误提示
try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D  # 注意：必须是 XDRPositiveSize2D
    from openpyxl.utils import column_index_from_string  # 如后续想通用列字母→索引可用

except Exception:
    Workbook = None
    XLImage = None

try:
    from PIL import Image
except Exception:
    Image = None


BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def _abs_path_from_web(rel: str) -> str:
    """
    将前端保存的“web 路径”转为真实磁盘路径。
    支持两类：
      1) /files/system/...   → 映射到 SYSTEM_IMAGE_ROOT（网络盘）
      2) 其它（/static/...） → 仍按项目根目录拼接
    """
    from urllib.parse import urlsplit

    r = (rel or "")
    # 去掉 query/hash（例如 ?v=xxx 的防缓存参数）
    parsed = urlsplit(r)
    r = parsed.path or r

    r = r.replace("\\", "/")

    # 绝对 URL 或 data: 不处理为本地文件，原样返回，调用方自行跳过
    if r.startswith("http://") or r.startswith("https://") or r.startswith("data:"):
        return r

    # /files/system/... → SYSTEM_IMAGE_ROOT 下的真实文件
    if r.startswith("/files/system/") or r.startswith("files/system/"):
        # 截出 /files/system/ 之后的相对部分
        sub = r.split("/files/system/", 1)[-1] if "/files/system/" in r else r.split("files/system/", 1)[-1]
        return _safe_join_system_root(sub)

    # 默认：按项目根目录拼接（兼容 /static/uploads/... 等历史路径）
    rel2 = r.lstrip("/").replace("/", os.sep)
    return os.path.join(BASE_DIR, rel2)


# === System 图片只读映射：将 /files/system/... 映射到 SYSTEM_IMAGE_ROOT 物理路径 ===
from config import SYSTEM_IMAGE_ROOT


def _safe_join_system_root(subpath: str) -> str:
    """
    将 web 子路径安全拼接到 SYSTEM_IMAGE_ROOT，防止路径穿越。
    例如 subpath = "2024/2408/240824_A/250824_A_1.jpg"
    """
    subpath = (subpath or "").replace("\\", "/").lstrip("/")
    parts = [p for p in subpath.split("/") if p not in ("", ".", "..")]
    full = os.path.join(SYSTEM_IMAGE_ROOT, *parts)
    # 安全校验：确保仍在 SYSTEM_IMAGE_ROOT 内
    root_norm = os.path.abspath(SYSTEM_IMAGE_ROOT)
    full_norm = os.path.abspath(full)
    if not full_norm.startswith(root_norm):
        raise ValueError("非法路径")
    return full


# =============================== 常量与目录（上传路径） ===============================
UPLOAD_ROOT = os.path.join("static", "uploads", "items")
os.makedirs(UPLOAD_ROOT, exist_ok=True)
ALLOW_UPSCALE = True
EMU_PER_PX = 9525


# =============================== 工具：Excel风格字母序转数字 ===============================
def code_to_number(code: str) -> int:
    """
    Excel 风格字母序 → 数字：A=1..Z=26, AA=27..
    - 用于对出品人 seller_code 做“自然序”的排序
    """
    if not code:
        return 0
    n = 0
    for ch in str(code).strip().upper():
        if not ('A' <= ch <= 'Z'):
            return 0
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n


# ===== 内部编号自然排序（统一入口） =====
import re

_NAT_RE = re.compile(r"^(.*?)(?:[_-]?(\d+))?$")


def item_code_nat_key_from_code(code: str):
    """将 '250822_BB_12' -> ('250822_BB', 12)；无数字则返回 0。"""
    s = str(code or "")
    m = _NAT_RE.match(s)
    if not m:
        return (s, 0)
    prefix, num = m.group(1), m.group(2)
    return (prefix, int(num) if num is not None else 0)


def sort_items_by_code(items):
    """
    统一对 items 排序（支持字典或 ORM 对象）
    - 字典：读取 item_code 字段
    - 对象：读取 .item_code 属性
    """

    def _get_code(x):
        return x.get("item_code") if isinstance(x, dict) else getattr(x, "item_code", None)

    return sorted(items, key=lambda x: item_code_nat_key_from_code(_get_code(x)))


def pt_to_px(pt: float) -> float:
    return pt * 96.0 / 72.0


def colwidth_to_px(w: float) -> int:
    return int(round(7.0 * float(w) + 5.0))


def get_cell_box_px(ws, col_letter: str, row_idx: int):
    """取得某单元格的像素盒子（列宽×行高，单位：px）"""
    cw_chars = ws.column_dimensions[col_letter].width or 8.43
    cell_w_px = colwidth_to_px(cw_chars)
    rh_pt = ws.row_dimensions[row_idx].height
    if rh_pt is None:
        rh_pt = 15  # Excel 默认行高 pt
    cell_h_px = int(round(pt_to_px(rh_pt)))
    return cell_w_px, cell_h_px

def convert_excel_to_pdf(excel_path: Path, pdf_path: Path, *, open_visible: bool = False) -> None:
    """
    使用 Excel 的 ExportAsFixedFormat 严格按打印预览导出为 PDF。
    :param excel_path: 需要导出的 Excel 文件路径（.xlsx / .xlsm / .xls 等）
    :param pdf_path: 输出的 PDF 完整路径（若父目录不存在会创建）
    :param open_visible: 是否显示 Excel 窗口（默认不显示）
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"找不到 Excel 文件：{excel_path}")

    pdf_path.parent.mkdir(parents=True, exist_ok=True)

    # 常量：Excel 中的类型 0=PDF
    xlTypePDF = 0
    # 取整个工作簿（0）或选中区域（3）/指定页等，通常导出整个工作簿最贴近你的要求
    xlQualityStandard = 0  # 标准质量
    include_doc_props = True
    ignore_print_areas = False  # 非常关键：False 时严格按你在每个工作表设置的打印区域导出
    open_after_publish = False

    # 初始化 COM
    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = open_visible
        excel.DisplayAlerts = False

        # 打开不更新外部链接，避免弹窗阻塞
        workbook = excel.Workbooks.Open(
            str(excel_path),
            UpdateLinks=0,      # 不更新外部链接
            ReadOnly=True,      # 以只读打开，避免被占用
            IgnoreReadOnlyRecommended=True
        )

        # 关键：按工作簿的打印设置导出为 PDF
        # 如果你只想导出某些工作表，可以在这里选择 Sheets([...]).Select() 再调用 ActiveSheet.ExportAsFixedFormat。
        workbook.ExportAsFixedFormat(
            Type=xlTypePDF,
            Filename=str(pdf_path),
            Quality=xlQualityStandard,
            IncludeDocProperties=include_doc_props,
            IgnorePrintAreas=ignore_print_areas,
            OpenAfterPublish=open_after_publish
        )

    finally:
        try:
            if workbook is not None:
                workbook.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()

    if not pdf_path.exists():
        raise RuntimeError("导出失败：未在预期位置生成 PDF。")



# =============================== Flask App Factory ===============================
def create_app():
    app = Flask(__name__, static_folder="static", template_folder="templates")
    app.config.from_object("config")

    # === API 错误统一 JSON：避免前端 .json() 解析到 HTML 报错 ===
    @app.errorhandler(404)
    def _api_404(e):
        if request.path.startswith("/api/"):
            return jsonify({"error": "not found"}), 404
        return e

    @app.errorhandler(405)
    def _api_405(e):
        if request.path.startswith("/api/"):
            return jsonify({"error": "method not allowed"}), 405
        return e

    @app.errorhandler(500)
    def _api_500(e):
        if request.path.startswith("/api/"):
            # 不把 HTML 500 页返回给前端
            return jsonify({"error": "internal server error"}), 500
        return e

    # ============================================================================
    # 一、页面视图（HTML 渲染）
    # ============================================================================

    @app.route("/files/system/<path:subpath>")
    def serve_system_file(subpath):
        """
        只读文件服务：将 UNC 根目录下的文件通过 HTTP 提供给前端 <img src> 直接使用。
        """
        try:
            full = _safe_join_system_root(subpath)
        except Exception:
            abort(400)

        if not os.path.isfile(full):
            abort(404)
        # as_attachment=False 表示内联显示
        return send_file(full, as_attachment=False)

    # === 新增：系统图片缩略图（列表小图用） ===
    @app.route("/thumb/system/<int:size>/<path:subpath>")
    def serve_system_thumb(size, subpath):
        """
        缩略图接口：
        /thumb/system/120/2024/2408/240824_A/xxx.jpg

        - size：最长边像素，限定在 40~400 之间
        - 仅对 /files/system/... 的图片有效
        - 若没安装 PIL，则回退到原图
        """
        # 如果没装 PIL，直接退回原图
        if Image is None:
            return serve_system_file(subpath)

        # 限制一下 size，避免太大
        size = max(40, min(int(size or 120), 400))

        try:
            full = _safe_join_system_root(subpath)
        except Exception:
            abort(400)

        if not os.path.isfile(full):
            abort(404)

        try:
            from io import BytesIO

            with Image.open(full) as im:
                # 转成 RGB，防止某些模式保存 JPEG 出问题
                im = im.convert("RGB")
                # 最长边 = size，等比缩放（contain）
                im.thumbnail((size, size))

                buf = BytesIO()
                im.save(buf, format="JPEG", quality=80)
                buf.seek(0)
            return send_file(buf, mimetype="image/jpeg")
        except Exception:
            # 出错时兜底返回原图
            return serve_system_file(subpath)

    # [HTML] 首页：templates/index.html
    @app.route("/")
    def index():
        return render_template("index.html")

    # [HTML] 在库管理 - 入库页：templates/inventory/inbound.html
    @app.route("/inventory/inbound")
    def inventory_inbound():
        return render_template("inventory/inbound.html")

    # [HTML] 在库管理 - 批次总览：templates/inventory/batches.html
    @app.route("/inventory/batches")
    def inventory_batches():
        return render_template("inventory/batches.html")

    # [HTML] 在库管理 - 指定批次编辑页：templates/inventory/batch_edit.html
    @app.route("/inventory/batch/<stockin_date>/<seller_code>")
    def inventory_batch_edit(stockin_date, seller_code):
        return render_template(
            "inventory/batch_edit.html",
            stockin_date=stockin_date,
            seller_code=seller_code
        )

    # [HTML] 在库管理 - 批次打印视图：templates/inventory/print_batch.html
    @app.route("/inventory/print/batch/<stockin_date>/<seller_code>")
    def inventory_print_batch(stockin_date, seller_code):
        session = get_session()
        try:
            items = _fetch_batch_items(session, stockin_date, seller_code)
            seller_name = _fetch_seller_name(session, seller_code)
            missing = _check_missing_images(items)
            items = sort_items_by_code(items)

            batch_code = _format_batch_code(stockin_date, seller_code)  # ← 新增

            return render_template(
                "inventory/print_batch.html",
                stockin_date=stockin_date,
                seller_code=seller_code,
                seller_name=seller_name,
                print_time=datetime.now().strftime("%Y-%m-%d %H:%M"),
                items=items,
                missing=missing,
                batch_code=batch_code  # ← 新增
            )
        finally:
            session.close()

    # [HTML] 在库管理 - 单件编辑：templates/inventory/item_edit.html
    @app.route("/inventory/item/<item_code>")
    def inventory_item_edit(item_code):
        # 从 query 里接收批次上下文，传给模板（用于上一件/下一件和返回）
        stockin_date = request.args.get("stockin_date")
        seller_code = request.args.get("seller_code")
        return render_template(
            "inventory/item_edit.html",
            item_code=item_code,
            stockin_date=stockin_date,
            seller_code=seller_code
        )

    # [HTML] 在库管理 - 在库查询/修改：templates/inventory/list.html
    @app.route("/inventory/list")
    def inventory_list():
        return render_template("inventory/list.html")

    # [HTML] 在库管理 - 批次总览（别名路径）
    @app.route("/batches")
    def batches_index():
        return render_template("inventory/batches.html")

    # [HTML] 出库页面（不上拍出库）
    @app.route("/inventory/outbound")
    def inventory_outbound():
        return render_template("inventory/outbound.html")

    # [HTML] 拍卖会管理 - 制作：templates/auctions/build.html
    @app.route("/auctions/build")
    def auctions_build():
        return render_template("auctions/build.html")

    # [HTML] 拍卖会管理 - 日常图片管理：templates/auctions/images.html
    @app.route("/auctions/images")
    def auctions_images():
        return render_template("auctions/images.html")

    # [HTML] 拍卖会管理 - 账单制作：templates/auctions/billing.html
    @app.route("/auctions/billing")
    def auctions_billing():
        return render_template("auctions/billing.html")

    # [HTML] 拍卖会管理 - 返品单制作：templates/auctions/returns.html
    @app.route("/auctions/returns")
    def auctions_returns():
        return render_template("auctions/returns.html")

    # [HTML] 出品人 - 列表页：templates/sellers/index.html
    @app.route("/sellers")
    def sellers_index():
        return render_template("sellers/index.html")

    # [HTML] 出品人 - 新增页：templates/sellers/new.html
    @app.route("/sellers/new")
    def sellers_new():
        return render_template("sellers/new.html")

    # [HTML] 买方 - 列表页：templates/buyers/index.html
    @app.route("/buyers")
    def buyers_index():
        return render_template("buyers/index.html")

    # [HTML] 入库设置页（聚合入口）：templates/settings/stockin_settings.html
    @app.route("/settings/stockin", methods=["GET"])
    def settings_stockin():
        return render_template("settings/stockin_settings.html", title="入库设置")

    # [HTML] 设置 - 物品种类：templates/settings/item_categories.html
    @app.route("/settings/item-categories", methods=["GET"])
    def settings_item_categories():
        return render_template("settings/item_categories.html", title="物品种类设置")

    # [HTML] 设置 - 物品状态：templates/settings/item_statuses.html
    @app.route("/settings/item_statuses", methods=["GET"])
    def settings_item_statuses_page():
        return render_template("settings/item_statuses.html", title="物品状态设置")

    # [API] 设置 - 材质枚举（colors / materials / shapes），仅 GET
    @app.route("/api/settings/material_options", methods=["GET"])
    def api_settings_material_options():
        session = get_session()
        try:
            rows = session.query(MaterialOption).filter(MaterialOption.enabled == 1).all()
            out = {"颜色": [], "材质": [], "形制": []}
            for r in rows:
                g = (r.group_name or "").strip()
                if g in out:
                    out[g].append(r.name)
            # 去重+排序（可选）
            for k in out:
                out[k] = sorted(list({x.strip() for x in out[k] if x and x.strip()}))
            return jsonify(out)
        finally:
            session.close()

    # [HTML] 设置 - 箱号：templates/settings/boxes.html
    @app.route("/settings/boxes", methods=["GET"])
    def settings_boxes_page():
        return render_template("settings/boxes.html", title="箱号设置")

    # [HTML] 设置 - 附属品：templates/settings/accessory_types.html
    @app.route("/settings/accessory-types", methods=["GET"])
    def settings_accessory_types_page():
        return render_template("settings/accessory_types.html", title="附属品设置")

    # ==================== 拖拽排序 API（新增，不影响原有 CRUD） ====================
    from sqlalchemy import text

    # [HTML] 设置 - 材质选项管理页
    @app.route("/settings/material-options", methods=["GET"])
    def settings_material_options_page():
        return render_template("settings/material_options.html", title="材质选项")

    # ---- Boxes ----
    @app.route("/api/settings/boxes/list", methods=["GET"])
    def api_boxes_list():
        session = get_session()
        try:
            rows = session.execute(text(
                "SELECT box_code AS name, COALESCE(sort,0) AS sort FROM boxes ORDER BY sort ASC, box_code ASC")).fetchall()
            return jsonify({"items": [{"name": r.name, "sort": int(r.sort)} for r in rows]})
        finally:
            session.close()

    @app.route("/api/settings/boxes/reorder", methods=["POST"])
    def api_boxes_reorder():
        session = get_session()
        try:
            data = request.get_json(silent=True) or {}
            names = data.get("names") or []
            for i, name in enumerate(names, start=1):
                session.execute(text("UPDATE boxes SET sort=:s WHERE box_code=:n"), {"s": i, "n": name})
            session.commit()
            return jsonify({"ok": True, "count": len(names)})
        except Exception as e:
            session.rollback()
            return jsonify({"ok": False, "error": str(e)}), 400
        finally:
            session.close()

    # ---- Accessory Types ----
    @app.route("/api/settings/accessory_types/list", methods=["GET"])
    def api_acc_list():
        session = get_session()
        try:
            rows = session.execute(text(
                "SELECT accessory_name AS name, COALESCE(sort,0) AS sort FROM accessory_types ORDER BY sort ASC, accessory_name ASC")).fetchall()
            return jsonify({"items": [{"name": r.name, "sort": int(r.sort)} for r in rows]})
        finally:
            session.close()

    @app.route("/api/settings/accessory_types/reorder", methods=["POST"])
    def api_acc_reorder():
        session = get_session()
        try:
            data = request.get_json(silent=True) or {}
            names = data.get("names") or []
            for i, name in enumerate(names, start=1):
                session.execute(text("UPDATE accessory_types SET sort=:s WHERE accessory_name=:n"), {"s": i, "n": name})
            session.commit()
            return jsonify({"ok": True, "count": len(names)})
        except Exception as e:
            session.rollback()
            return jsonify({"ok": False, "error": str(e)}), 400
        finally:
            session.close()

    # ---- Item Statuses ----
    @app.route("/api/settings/item_statuses/list", methods=["GET"])
    def api_status_list():
        session = get_session()
        try:
            group = (request.args.get("group") or "").strip()
            if group:
                q = text(
                    "SELECT item_status AS name, group_name, COALESCE(sort,0) AS sort FROM item_statuses WHERE COALESCE(group_name,'')=:g ORDER BY sort ASC, item_status ASC")
                rows = session.execute(q, {"g": group}).fetchall()
            else:
                q = text(
                    "SELECT item_status AS name, group_name, COALESCE(sort,0) AS sort FROM item_statuses ORDER BY sort ASC, item_status ASC")
                rows = session.execute(q).fetchall()
            return jsonify({"items": [{"name": r.name, "group_name": r.group_name, "sort": int(r.sort)} for r in rows]})
        finally:
            session.close()

    @app.route("/api/settings/item_statuses/reorder", methods=["POST"])
    def api_status_reorder():
        session = get_session()
        try:
            data = request.get_json(silent=True) or {}
            names = data.get("names") or []
            group = (data.get("group") or "").strip()
            for i, name in enumerate(names, start=1):
                if group:
                    session.execute(
                        text("UPDATE item_statuses SET sort=:s WHERE item_status=:n AND COALESCE(group_name,'')=:g"),
                        {"s": i, "n": name, "g": group})
                else:
                    session.execute(text("UPDATE item_statuses SET sort=:s WHERE item_status=:n"), {"s": i, "n": name})
            session.commit()
            return jsonify({"ok": True, "count": len(names)})
        except Exception as e:
            session.rollback()
            return jsonify({"ok": False, "error": str(e)}), 400
        finally:
            session.close()

    @app.route("/api/settings/item_statuses_add", methods=["POST"])
    def api_status_add():
        session = get_session()
        try:
            data = request.get_json(silent=True) or {}
            name = (data.get("name") or "").strip()
            group = (data.get("group") or "").strip() or None
            if not name:
                return jsonify({"ok": False, "error": "name required"}), 400
            if group:
                maxs = session.execute(
                    text("SELECT COALESCE(MAX(sort),0) FROM item_statuses WHERE COALESCE(group_name,'')=:g"),
                    {"g": group}).scalar()
            else:
                maxs = session.execute(text("SELECT COALESCE(MAX(sort),0) FROM item_statuses")).scalar()
            session.execute(text("INSERT INTO item_statuses(item_status, group_name, sort) VALUES(:n, :g, :s)"),
                            {"n": name, "g": group, "s": int(maxs) + 1})
            session.commit()
            return jsonify({"ok": True})
        except Exception as e:
            session.rollback()
            if "UNIQUE" in str(e).upper():
                return jsonify({"ok": False, "error": "duplicate"}), 409
            return jsonify({"ok": False, "error": str(e)}), 400
        finally:
            session.close()

    @app.route("/api/settings/item_statuses_delete", methods=["POST"])
    def api_status_delete():
        session = get_session()
        try:
            data = request.get_json(silent=True) or {}
            name = (data.get("name") or "").strip()
            if not name:
                return jsonify({"ok": False, "error": "name required"}), 400
            session.execute(text("DELETE FROM item_statuses WHERE item_status=:n"), {"n": name})
            session.commit()
            return jsonify({"ok": True})
        except Exception as e:
            session.rollback()
            return jsonify({"ok": False, "error": str(e)}), 400
        finally:
            session.close()

    # ---- Material Options ----
    @app.route("/api/settings/material_options/list", methods=["GET"])
    def api_mat_list():
        session = get_session()
        try:
            group = (request.args.get("group") or "").strip()
            if not group:
                return jsonify({"ok": False, "error": "group required"}), 400
            rows = session.execute(text(
                "SELECT name, COALESCE(sort,0) AS sort FROM material_options WHERE group_name=:g ORDER BY sort ASC, name ASC"),
                                   {"g": group}).fetchall()
            return jsonify({"items": [{"name": r.name, "sort": int(r.sort)} for r in rows]})
        finally:
            session.close()

    @app.route("/api/settings/material_options/reorder", methods=["POST"])
    def api_mat_reorder():
        session = get_session()
        try:
            data = request.get_json(silent=True) or {}
            group = (data.get("group") or "").strip()
            names = data.get("names") or []
            if not group:
                return jsonify({"ok": False, "error": "group required"}), 400
            for i, name in enumerate(names, start=1):
                session.execute(text("UPDATE material_options SET sort=:s WHERE group_name=:g AND name=:n"),
                                {"s": i, "g": group, "n": name})
            session.commit()
            return jsonify({"ok": True, "count": len(names)})
        except Exception as e:
            session.rollback()
            return jsonify({"ok": False, "error": str(e)}), 400
        finally:
            session.close()

    # 复用 GET /api/settings/material_options 给业务端；这里增加 POST/DELETE 供管理页使用
    @app.route("/api/settings/material_options", methods=["POST", "DELETE"])
    def api_mat_mutations():
        session = get_session()
        try:
            data = request.get_json(silent=True) or {}
            group = (data.get("group") or "").strip()
            name = (data.get("name") or "").strip()
            if not group or not name:
                return jsonify({"ok": False, "error": "group/name required"}), 400
            if request.method == "POST":
                maxs = session.execute(text("SELECT COALESCE(MAX(sort),0) FROM material_options WHERE group_name=:g"),
                                       {"g": group}).scalar()
                session.execute(text("INSERT INTO material_options(group_name, name, sort) VALUES(:g, :n, :s)"),
                                {"g": group, "n": name, "s": int(maxs) + 1})
                session.commit()
                return jsonify({"ok": True})
            else:
                session.execute(text("DELETE FROM material_options WHERE group_name=:g AND name=:n"),
                                {"g": group, "n": name})
                session.commit()
                return jsonify({"ok": True})
        except Exception as e:
            session.rollback()
            if "UNIQUE" in str(e).upper():
                return jsonify({"ok": False, "error": "duplicate"}), 409
            return jsonify({"ok": False, "error": str(e)}), 400
        finally:
            session.close()

    # ============================================================================
    # 二、系统健康探针 / 版本 / DB 连通性
    # ============================================================================

    # [API] 健康检查
    @app.route("/api/health")
    def api_health():
        return jsonify({"status": "ok"})

    # [API] 版本信息
    @app.route("/api/version")
    def api_version():
        return jsonify({"app": "auction-system-skeleton", "version": "0.2.0-m1"})

    # [API] DB 连通性
    @app.route("/api/db-check")
    def api_db_check():
        try:
            session = get_session()
            items = session.query(Item).count()
            sellers = session.query(Seller).count()
            auctions = session.query(Auction).count()
            buyers = session.query(Buyer).count()
            session.close()
            return jsonify({"db": "ok", "items": items, "sellers": sellers, "auctions": auctions, "buyers": buyers})
        except Exception as e:
            return jsonify({"db": "error", "message": str(e)}), 500

    # ============================================================================
    # 三、基础设置（物品种类 / 箱号 / 附属品）—— API
    # ============================================================================

    # ---- 基础设置 · 通用工具：读取/新增/删除（简单单列表） ----
    def _list_simple_table(session, table, field):
        """按字典序列出简单表的单列值"""
        rs = session.execute(text(f"SELECT {field} FROM {table} ORDER BY {field} ASC")).fetchall()
        return [r[0] for r in rs]

    def _add_simple_value(session, table, field, value):
        """插入简单表的单列值（去重）"""
        if not value or not str(value).strip():
            raise ValueError("值不能为空")
        v = str(value).strip()
        exist = session.execute(text(f"SELECT COUNT(*) FROM {table} WHERE {field} = :v"), {"v": v}).scalar()
        if exist and int(exist) > 0:
            return False
        session.execute(text(f"INSERT INTO {table}({field}) VALUES(:v)"), {"v": v})
        return True

    def _delete_simple_value(session, table, field, value):
        """删除简单表的单列值"""
        v = str(value).strip()
        session.execute(text(f"DELETE FROM {table} WHERE {field} = :v"), {"v": v})
        return True

    # [API] 设置 - 箱号
    @app.route("/api/settings/boxes", methods=["GET", "POST", "PUT", "DELETE"])
    def api_settings_boxes():
        session = get_session()
        try:
            table, field = "boxes", "box_code"
            if request.method == "GET":
                rs = session.execute(text(f"""
                    SELECT {field} FROM {table}
                    ORDER BY COALESCE(sort, 1000000000) ASC, {field} ASC
                """)).fetchall()
                return jsonify({"items": [r[0] for r in rs]})

            elif request.method == "POST":
                data = request.get_json(silent=True) or {}
                ok = _add_simple_value(session, table, field, data.get("name"))
                session.commit()
                return jsonify({"ok": True, "created": ok})

            elif request.method == "PUT":
                data = request.get_json(silent=True) or {}
                # 兼容多种键名：前端可能传 old_name/new_name 或 old/new，甚至只传 name 当作新值
                old_v = (data.get("old") or data.get("old_name") or "").strip()
                new_v = (data.get("new") or data.get("new_name") or data.get("name") or "").strip()
                if not old_v or not new_v:
                    session.rollback()
                    return jsonify({"ok": False, "error": "值不能为空"}), 400
                res = _update_simple_value(session, table, field, old_v, new_v)
                if res == "conflict":
                    session.rollback()
                    return jsonify({"ok": False, "error": "duplicate"}), 409
                session.commit()
                return jsonify({"ok": True, "updated": bool(res)})

            else:  # DELETE
                data = request.get_json(silent=True) or {}
                _delete_simple_value(session, table, field, data.get("name"))
                session.commit()
                return jsonify({"ok": True})

        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 400
        finally:
            session.close()

    # [API] 设置 - 附属品
    @app.route("/api/settings/accessory_types", methods=["GET", "POST", "PUT", "DELETE"])
    def api_settings_accessory_types():
        session = get_session()
        try:
            table, field = "accessory_types", "accessory_name"

            if request.method == "GET":
                rs = session.execute(text(f"""
                    SELECT {field} FROM {table}
                    ORDER BY COALESCE(sort, 1000000000) ASC, {field} ASC
                """)).fetchall()
                return jsonify({"items": [r[0] for r in rs]})

            elif request.method == "POST":
                data = request.get_json(silent=True) or {}
                ok = _add_simple_value(session, table, field, data.get("name"))
                session.commit()
                return jsonify({"ok": True, "created": ok})

            elif request.method == "PUT":
                data = request.get_json(silent=True) or {}
                # 兼容多种键名：前端可能传 old_name/new_name 或 old/new，甚至只传 name 当作新值
                old_v = (data.get("old") or data.get("old_name") or "").strip()
                new_v = (data.get("new") or data.get("new_name") or data.get("name") or "").strip()
                if not old_v or not new_v:
                    session.rollback()
                    return jsonify({"ok": False, "error": "值不能为空"}), 400
                res = _update_simple_value(session, table, field, old_v, new_v)
                if res == "conflict":
                    session.rollback()
                    return jsonify({"ok": False, "error": "duplicate"}), 409
                session.commit()
                return jsonify({"ok": True, "updated": bool(res)})

            else:  # DELETE
                data = request.get_json(silent=True) or {}
                _delete_simple_value(session, table, field, data.get("name"))
                session.commit()
                return jsonify({"ok": True})

        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 400
        finally:
            session.close()

    # ---- Item Categories（拖拽排序专用）----
    @app.route("/api/settings/item_categories/list", methods=["GET"])
    def api_item_categories_list():
        session = get_session()
        try:
            rows = session.execute(text("""
                SELECT item_category AS name, COALESCE(sort,0) AS sort
                FROM item_categories
                ORDER BY COALESCE(sort, 1000000000) ASC, item_category ASC
            """)).fetchall()
            return jsonify({"items": [{"name": r.name, "sort": int(r.sort)} for r in rows]})
        finally:
            session.close()

    @app.route("/api/settings/item_categories/reorder", methods=["POST"])
    def api_item_categories_reorder():
        session = get_session()
        try:
            data = request.get_json(silent=True) or {}
            names = data.get("names") or []
            for i, name in enumerate(names, start=1):
                session.execute(text(
                    "UPDATE item_categories SET sort=:s WHERE item_category=:n"
                ), {"s": i, "n": name})
            session.commit()
            return jsonify({"ok": True, "count": len(names)})
        except Exception as e:
            session.rollback()
            return jsonify({"ok": False, "error": str(e)}), 400
        finally:
            session.close()

    # [API] 设置 - 物品种类
    @app.route("/api/settings/item_categories", methods=["GET", "POST", "PUT", "DELETE"])
    def api_settings_item_categories():
        session = get_session()
        try:
            table, field = "item_categories", "item_category"
            if request.method == "GET":
                rs = session.execute(text("""
                    SELECT item_category
                    FROM item_categories
                    ORDER BY COALESCE(sort, 1000000000) ASC, item_category ASC
                """)).fetchall()
                return jsonify({"items": [r[0] for r in rs]})

            elif request.method == "POST":
                data = request.get_json(silent=True) or {}
                ok = _add_simple_value(session, table, field, data.get("name"))
                session.commit()
                return jsonify({"ok": True, "created": ok})

            elif request.method == "PUT":
                data = request.get_json(silent=True) or {}
                old_v = (data.get("old") or data.get("old_name") or "").strip()
                new_v = (data.get("new") or data.get("new_name") or data.get("name") or "").strip()
                if not old_v or not new_v:
                    session.rollback()
                    return jsonify({"ok": False, "error": "值不能为空"}), 400
                res = _update_simple_value(session, table, field, old_v, new_v)
                if res == "conflict":
                    session.rollback()
                    return jsonify({"ok": False, "error": "duplicate"}), 409
                session.commit()
                return jsonify({"ok": True, "updated": bool(res)})

            else:  # DELETE
                data = request.get_json(silent=True) or {}
                _delete_simple_value(session, table, field, data.get("name"))
                session.commit()
                return jsonify({"ok": True})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 400
        finally:
            session.close()

    # [API] 设置 - 物品状态
    @app.route("/api/settings/item_statuses", methods=["GET", "POST", "PUT", "DELETE"])
    def api_settings_item_statuses():
        session = get_session()
        try:
            table, field = "item_statuses", "item_status"
            if request.method == "GET":
                return jsonify({"items": _list_simple_table(session, table, field)})

            elif request.method == "POST":
                data = request.get_json(silent=True) or {}
                ok = _add_simple_value(session, table, field, data.get("name"))
                session.commit()
                return jsonify({"ok": True, "created": ok})

            elif request.method == "PUT":
                data = request.get_json(silent=True) or {}
                # 兼容多种键名：前端可能传 old_name/new_name 或 old/new，甚至只传 name 当作新值
                old_v = (data.get("old") or data.get("old_name") or "").strip()
                new_v = (data.get("new") or data.get("new_name") or data.get("name") or "").strip()
                if not old_v or not new_v:
                    session.rollback()
                    return jsonify({"ok": False, "error": "值不能为空"}), 400
                res = _update_simple_value(session, table, field, old_v, new_v)
                if res == "conflict":
                    session.rollback()
                    return jsonify({"ok": False, "error": "duplicate"}), 409
                session.commit()
                return jsonify({"ok": True, "updated": bool(res)})

            else:  # DELETE
                data = request.get_json(silent=True) or {}
                _delete_simple_value(session, table, field, data.get("name"))
                session.commit()
                return jsonify({"ok": True})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 400
        finally:
            session.close()

    # ============================================================================
    # 四、通用工具函数
    # ============================================================================

    def to_decimal_or_none(v):
        """小数安全转换：None/空 返回 None；其他转 Decimal，非法抛错"""
        if v in (None, "", "null", "None"):
            return None
        try:
            return Decimal(str(v))
        except InvalidOperation:
            raise ValueError("金额/数字字段格式错误")

    def to_int_or_none(v):
        """金额（万日元）整数化：None/空 -> None；其余必须为非负整数"""
        if v in (None, "", "null", "None"):
            return None
        s = str(v).strip()
        # 允许纯数字；禁止小数与负数
        if not re.fullmatch(r"\d+", s):
            raise ValueError("金额需为非负整数（万日元）")
        return int(s)

    def ensure_status(session, status_str):
        """校验物品状态在 ItemStatus 表中存在（返回合法值或 None）"""
        if not status_str:
            return None
        from create_database import ItemStatus
        s = session.get(ItemStatus, status_str)
        if not s:
            raise ValueError(f"无效物品状态：{status_str}")
        return status_str

    def log_op(session, entity_type, entity_id, action, before=None, after=None, operator="admin"):
        """通用操作日志"""
        op = OperationLog(
            entity_type=entity_type,
            entity_id=str(entity_id),
            action=action,
            before_value=before,
            after_value=after,
            operator=operator
        )
        session.add(op)

    def normalize_pct(value):
        """
        百分比输入归一化：
        - 接受 12 / "13.2%" / 0.132 等，统一换算为 0~1
        """
        if value is None or value == "":
            return None
        # 去掉可能的百分号
        if isinstance(value, str):
            v = value.strip()
            if v.endswith("%"):
                v = v[:-1].strip()
        else:
            v = value
        try:
            num = float(v)
        except Exception:
            raise ValueError("百分比格式错误（仅支持数字或带%）")

        if 0 <= num <= 1:
            return num
        if 0 <= num <= 100:
            return num / 100.0
        raise ValueError("百分比应在 0~100 之间")

    def _update_simple_value(session, table, field, old_value, new_value):
        """
        简表“改名”：把 table.field = old_value 的记录 改为 new_value
        - 若 new_value 已存在返回 'conflict'
        """
        if new_value is None:
            raise ValueError("值不能为空")
        v_old = str(old_value).strip()
        v_new = str(new_value).strip()
        if not v_new:
            raise ValueError("值不能为空")
        if v_old == v_new:
            return True  # 视为成功，无变化

        # 重名检查
        exist = session.execute(
            text(f"SELECT COUNT(*) FROM {table} WHERE {field} = :v"),
            {"v": v_new}
        ).scalar()
        if int(exist or 0) > 0:
            return "conflict"

        session.execute(
            text(f"UPDATE {table} SET {field} = :new WHERE {field} = :old"),
            {"new": v_new, "old": v_old}
        )
        return True

    # ---- Excel 列名风格 ↔ 数字（给出品人分配下一编码用） ----
    def code_to_num(code: str) -> int:
        code = (code or "").strip().upper()
        if not code.isalpha():
            return 0
        n = 0
        for ch in code:
            n = n * 26 + (ord(ch) - ord('A') + 1)
        return n

    def num_to_code(n: int) -> str:
        if n <= 0:
            return "A"
        s = []
        while n > 0:
            n -= 1
            s.append(chr(n % 26 + ord('A')))
            n //= 26
        return "".join(reversed(s))

    # ============================================================================
    # 五、Items（物品）相关接口
    # ============================================================================

    # [API] 生成下一个出品人编码（Excel 序）
    @app.route("/api/sellers/next-code")
    def api_sellers_next_code():
        session = get_session()
        try:
            codes = [s.seller_code for s in session.query(Seller.seller_code).all() if s.seller_code]
            max_num = 0
            for c in codes:
                max_num = max(max_num, code_to_num(c))
            next_code = num_to_code(max_num + 1)
            return jsonify({"next_code": next_code})
        finally:
            session.close()

    # [API] 批量更新物品
    @app.route("/api/items/bulk-update", methods=["POST"])
    def api_items_bulk_update():
        """
        批量更新：
          - 普通字段：item_name / item_size / item_location / item_category / item_image
          - 附属品：通过映射表 item_accessories 覆盖保存（支持 accessories: [] 或 item_accessories: "盒、签"）
        """
        payload = request.get_json(silent=True) or {}
        items = payload.get("items") or []
        if not isinstance(items, list) or not items:
            return jsonify({"error": "items 不能为空"}), 400

        session = get_session()
        try:
            acc_updates = {}  # { item_code: ["盒","签"] }

            for row in items:
                code = (row.get("item_code") or "").strip()
                if not code:
                    continue

                it = session.get(Item, code)
                if not it:
                    continue

                # 记录更新前（仅普通字段）
                before = {
                    "item_name": it.item_name,
                    "item_size": it.item_size,
                    "item_location": it.item_location,
                    "item_category": it.item_category,
                    "item_image": it.item_image,
                    "item_material": it.item_material,
                }

                # 更新普通字段（空串 -> None；允许清空）
                changed = {}
                for k in ["item_name", "item_size", "item_location", "item_category", "item_image", "item_box_code",
                          "item_material"]:
                    if k in row:
                        v = row.get(k)
                        setattr(it, k, (v if v not in ("", None) else None))
                        changed[k] = v

                # —— 新增：起拍价 / 底价（万日元整数）——
                # 先读旧值
                sp = it.starting_price
                rp = it.reserve_price
                if "starting_price" in row:
                    sp = to_int_or_none(row.get("starting_price"))  # None 允许清空
                if "reserve_price" in row:
                    rp = to_int_or_none(row.get("reserve_price"))  # None 允许清空
                if sp is not None and rp is not None and sp > rp:
                    raise ValueError("起拍价不能高于底价")
                it.starting_price = sp
                it.reserve_price = rp

                # 解析附属品
                acc_input = row.get("accessories", None)
                if acc_input is None and "item_accessories" in row:
                    acc_input = row.get("item_accessories")

                if acc_input is not None:
                    if isinstance(acc_input, str):
                        acc_list = [s.strip() for s in acc_input.replace(",", "、").split("、") if s.strip()]
                    else:
                        acc_list = [str(s).strip() for s in (acc_input or []) if str(s).strip()]
                    acc_updates[code] = acc_list

                # 日志
                try:
                    log_op(session, "item", code, "bulk_update",
                           before=str(before),
                           after=str(
                               {**changed, **({"accessories": acc_updates.get(code)} if code in acc_updates else {})}))
                except Exception:
                    pass

            session.flush()

            # 应用附属品（覆盖：先删后插）
            for code, acc_list in acc_updates.items():
                session.execute(text("DELETE FROM item_accessories WHERE item_code = :c"), {"c": code})
                for name in acc_list:
                    session.execute(
                        text("INSERT INTO item_accessories(item_code, accessory_name) VALUES(:c, :n)"),
                        {"c": code, "n": name}
                    )

            session.commit()
            return jsonify({"ok": True})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # [API] 批次内物品
    @app.route("/api/items/by-batch", methods=["GET"])
    def api_items_by_batch():
        session = get_session()
        try:
            stockin_date = request.args.get("stockin_date")
            seller_code = request.args.get("seller_code")

            # 兼容 batch_code=YYMMDD_A 或 YYYY-MM-DD_A
            if (not stockin_date or not seller_code):
                bc = (request.args.get("batch_code") or "").strip()
                if bc:
                    import re
                    m = re.match(r"^(\d{4}-\d{2}-\d{2})[_-]([A-Za-z]+)$", bc)
                    if m:
                        stockin_date = m.group(1)
                        seller_code = m.group(2).upper()
                    else:
                        m = re.match(r"^(\d{2})(\d{2})(\d{2})[_-]([A-Za-z]+)$", bc)
                        if m:
                            yyyy = str(2000 + int(m.group(1)))
                            stockin_date = f"{yyyy}-{m.group(2)}-{m.group(3)}"
                            seller_code = m.group(4).upper()

            if not stockin_date or not seller_code:
                return jsonify({"error": "缺少 stockin_date 或 seller_code"}), 400

            # 1) 批次内 items（只取必要字段）
            items = session.execute(text("""
                SELECT i.item_code,
                       i.item_name,
                       i.item_author,
                       i.item_status,
                       i.starting_price,
                       i.reserve_price,
                       i.item_size,  
                       i.item_location,
                       i.item_box_code,
                       i.item_category,
                       i.item_image,
                       i.stockin_date,
                       i.seller_code
                FROM items i
                WHERE i.stockin_date = :d AND i.seller_code = :s
                ORDER BY i.item_code
            """), {"d": stockin_date, "s": seller_code}).mappings().all()

            # 2) 附属品聚合
            acc_rows = session.execute(text("""
                SELECT ia.item_code, ia.accessory_name
                FROM item_accessories ia
                WHERE ia.item_code IN (
                    SELECT i2.item_code FROM items i2
                    WHERE i2.stockin_date = :d AND i2.seller_code = :s
                )
                ORDER BY ia.item_code, ia.accessory_name
            """), {"d": stockin_date, "s": seller_code}).fetchall()

            acc_map = {}
            for code, name in acc_rows:
                acc_map.setdefault(code, []).append(name)

            def to_row(m):
                code = m["item_code"]
                accessories = acc_map.get(code, [])
                return {
                    "item_code": code,
                    "item_name": m["item_name"],
                    "item_author": m["item_author"],
                    "item_status": m["item_status"],
                    "starting_price": m["starting_price"],
                    "reserve_price": m["reserve_price"],
                    "item_location": m["item_location"],
                    "item_box_code": m["item_box_code"],
                    "item_category": m.get("item_category"),
                    "item_image": m.get("item_image"),
                    "item_size": m.get("item_size"),
                    "stockin_date": m["stockin_date"],
                    "seller_code": m["seller_code"],
                    "accessories": accessories,
                    "accessories_text": "、".join(accessories)
                }

            rows = [to_row(m) for m in items]
            rows = sort_items_by_code(rows)  # ← 新增统一排序
            return jsonify({"total": len(rows), "items": rows})

        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # [API] 出库确认（当前仅实现「不上拍出库」）
    @app.route("/api/outbound/confirm", methods=["POST"])
    def api_outbound_confirm():
        """
        请求体格式：
        {
            "item_codes": ["240101_A_001", ...],
            "method": "post"   # "post" = 邮寄出库, "pickup" = 自提出库
            "date": "2025-11-27",
            "ship_type": "EMS / 佐川 / 黑猫 / 顺丰 等（可选，邮寄必填）"
        }
        """
        payload = request.json or {}
        item_codes = payload.get("item_codes") or []
        method = (payload.get("method") or "").strip()
        out_date = (payload.get("date") or "").strip()
        ship_type = (payload.get("ship_type") or "").strip()

        if not item_codes:
            return jsonify({"error": "缺少 item_codes"}), 400
        if method not in ("post", "pickup"):
            return jsonify({"error": "出库方式必须为 post 或 pickup"}), 400
        if not out_date:
            return jsonify({"error": "出库日期不能为空"}), 400
        if method == "post" and not ship_type:
            return jsonify({"error": "邮寄出库时必须填写邮寄种类"}), 400

        # 解析出库日期为 date 对象
        try:
            outbound_date_obj = datetime.strptime(out_date, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"error": "出库日期格式必须为 YYYY-MM-DD"}), 400

        # 操作人（暂时写死，后面接入登录后可改）
        operator = "admin"

        session = get_session()

        try:
            from create_database import Item

            # 当前版本：只处理「不上拍出库」，直接统一改成：
            #   - 邮寄出库 => 不上拍已寄出
            #   - 自提出库 => 不上拍已提货
            # 将来扩展成交品 / 返品时，可以在这里按业务类型选择不同状态
            if method == "post":
                target_status = "不上拍已寄出"
            else:
                target_status = "不上拍已提货"

            updated = 0
            for code in item_codes:
                it = session.get(Item, code)
                if not it:
                    continue

                before = {"item_status": it.item_status}
                it.item_status = target_status

                # 记录操作日志（包含出库方式、日期、邮寄种类）
                try:
                    log_op(
                        session,
                        "item",
                        code,
                        "outbound_confirm",
                        before=str(before),
                        after=str({
                            "item_status": target_status,
                            "outbound_method": method,
                            "outbound_date": out_date,
                            "ship_type": ship_type
                        })
                    )
                except Exception:
                    # 日志失败不影响主流程
                    pass

                # 新增：记录出库日志（结构化信息）
                # - outbound_date: 出库日期（前端输入）
                # - recorded_at: 录入日期时间（系统当前）
                # - seller_name / seller_code: 出品人
                # - item_code: 物品编号
                # - auction_id / lot_number: 场次 & LOT（拍卖会功能完成后可补齐，目前先 None）
                # - operator: 操作人
                # - method: 出库方式（post / pickup）
                # - ship_type: 快递种类（仅邮寄时有效）
                log_payload = {
                    "outbound_date": out_date,
                    "recorded_at": datetime.utcnow().isoformat(timespec="seconds"),
                    "seller_name": it.seller_name,
                    "seller_code": it.seller_code,
                    "item_code": code,
                    "auction_id": None,
                    "lot_number": None,
                    "operator": operator,
                    "method": method,
                    "ship_type": ship_type if method == "post" else ""
                }

                outbound_log = OutboundLog(
                    item_code=code,
                    outbound_type="other",  # 不上拍直接出库
                    ref_id=None,  # 将来有“出库单号”等再用
                    outbound_date=outbound_date_obj,
                    handled_by=operator,
                    notes=json.dumps(log_payload, ensure_ascii=False)
                )
                session.add(outbound_log)

                updated += 1

            session.commit()
            return jsonify({"ok": True, "updated": updated, "target_status": target_status})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # [API] 图片上传：强制用“编号+原后缀”命名，按 item_code 分目录保存
    @app.route("/api/upload-image", methods=["POST"])
    def api_upload_image():
        if "file" not in request.files:
            return jsonify({"error": "未选择文件"}), 400
        file = request.files["file"]
        item_code = (request.form.get("item_code") or "").strip()
        if not item_code:
            return jsonify({"error": "缺少 item_code"}), 400
        if file.filename == "":
            return jsonify({"error": "文件名为空"}), 400

        # 统一重命名：item_code + 原后缀（与前端一致，后端再确保一次）
        orig_name = file.filename
        _, ext = os.path.splitext(orig_name)
        ext = (ext or "").lower()
        # 常见图片类型兜底
        if not ext:
            mime = (file.mimetype or "").lower()
            mime_map = {"image/jpeg": ".jpg", "image/png": ".png", "image/webp": ".webp", "image/gif": ".gif"}
            ext = mime_map.get(mime, "")

        filename = secure_filename(f"{item_code}{ext}")

        # === 新逻辑：保存到 SYSTEM 路径：年（YYYY）/ 月（YYMM）/ 批次号（YYMMDD_S）/ 文件 ===
        # 1) 通过 item_code 反查该条目的 stockin_date、seller_code
        session = get_session()
        try:
            from create_database import Item
            it = session.get(Item, item_code)
            if not it:
                return jsonify({"error": "item 不存在"}), 404

            stockin_date = str(it.stockin_date)  # YYYY-MM-DD
            seller_code = it.seller_code

            # 2) 年、月（YYMM）、批次号（YYMMDD_S）
            from datetime import datetime
            dt = datetime.strptime(stockin_date, "%Y-%m-%d")
            year_folder = str(dt.year)  # e.g. "2024"
            yymm = dt.strftime("%y%m")  # e.g. "2408"
            batch_code = _format_batch_code(stockin_date, seller_code)  # e.g. "240824_A"

            # 3) 目录：\\...\\system\\{year}\\{yymm}\\{batch_code}
            dest_dir = os.path.join(SYSTEM_IMAGE_ROOT, year_folder, yymm, batch_code)
            os.makedirs(dest_dir, exist_ok=True)

            # 4) 文件名：内部编号（item_code）+ 原扩展名
            save_path = os.path.join(dest_dir, filename)

            # 5) 保存（同名覆盖）
            file.save(save_path)

            # 6) 返回给前端可直接 <img src> 的 Web 路径（由 serve_system_file 路由提供）
            rel_path = f"/files/system/{year_folder}/{yymm}/{batch_code}/{filename}"
            return jsonify({"ok": True, "path": rel_path})
        finally:
            session.close()

    # [API] 入库：按日期/出品人生成空白物品（批次）
    @app.route("/api/stock-batches/generate-items", methods=["POST"])
    def api_generate_items():
        from create_database import StockBatch, Item, Seller, ItemStatus
        payload = request.json or {}
        stockin_date = payload.get("stockin_date")  # "YYYY-MM-DD"
        seller_code = (payload.get("seller_code") or "").strip()
        count = int(payload.get("count") or 0)
        receiver = payload.get("stockin_receiver")
        staff = payload.get("stockin_staff")

        if not stockin_date or not seller_code or count <= 0:
            return jsonify({"error": "入库日期/出品人/件数 必填且有效"}), 400
        if not receiver:
            return jsonify({"error": "签收人必填"}), 400

        session = get_session()
        try:
            # 将字符串日期转为 date（SQLAlchemy Date 需要）
            from datetime import datetime
            try:
                dt = datetime.strptime(stockin_date, "%Y-%m-%d")
                stockin_date_obj = dt.date()
            except Exception:
                return jsonify({"error": "入库日期格式错误，必须为 YYYY-MM-DD"}), 400

            # 1) 校验出品人存在
            if not session.get(Seller, seller_code):
                return jsonify({"error": "出品人不存在，请先到出品人页面新增"}), 400

            # 2) 创建/确认批次
            sb = session.get(StockBatch, {"stockin_date": stockin_date_obj, "seller_code": seller_code})
            if not sb:
                sb = StockBatch(
                    stockin_date=stockin_date_obj,
                    seller_code=seller_code,
                    stockin_count=0,  # ← 先置 0，稍后按“实际新增件数”累加
                    has_physical_list=bool(payload.get("has_physical_list", False)),
                    stockin_receiver=receiver,
                    stockin_staff=staff,
                )
                session.add(sb)
            else:
                # 不在这里改 stockin_count，保持签收人等信息
                sb.stockin_receiver = receiver
                if staff:
                    sb.stockin_staff = staff

            # 3) 生成 item_code 并新建空白 item（若存在则跳过）
            dt = datetime.strptime(stockin_date, "%Y-%m-%d")
            prefix = dt.strftime("%y%m%d") + f"_{seller_code}_"
            created = 0
            created_codes = []
            for i in range(1, count + 1):
                code = f"{prefix}{i}"
                if not session.get(Item, code):
                    default_status = None
                    try:
                        from create_database import ItemStatus
                        # 优先使用「待上拍」作为默认状态
                        default_status = session.get(ItemStatus, "待上拍")
                    except Exception:
                        pass
                    it = Item(
                        item_code=code,
                        item_name=None, item_size=None, item_image=None,
                        is_in_box=False, item_box_code=None, item_location=None,
                        item_category=None, reserve_price=None, seller_name=None,
                        seller_code=seller_code, starting_price=None,
                        stockin_date=stockin_date_obj, item_order=None, item_barcode=None,
                        photo_date_shot=None, photo_date_detail=None, photo_date_ps=None,
                        item_material=None, item_seal=None, item_inscription=None,
                        item_description=None, item_author=None,
                        item_status=("待上拍" if default_status else None),
                        item_notes=None
                    )

                    session.add(it)
                    created += 1
                created_codes.append(code)
            # 统一：把“实际新增件数”累加到批次
            sb.stockin_count = (sb.stockin_count or 0) + created

            session.commit()
            return jsonify({"ok": True, "item_codes": created_codes})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # [API] 获取单件详情（含附属品）
    @app.route("/api/items/<item_code>", methods=["GET"])
    def api_items_get(item_code):
        session = get_session()
        try:
            it = session.get(Item, item_code)
            if not it:
                return jsonify({"error": "item 不存在"}), 404

            # 附属品列表
            acc_rows = session.execute(
                text("SELECT accessory_name FROM item_accessories WHERE item_code = :c"),
                {"c": item_code}
            ).fetchall()
            accessories = [r[0] for r in acc_rows]

            return jsonify({
                "item_code": it.item_code,
                "item_name": it.item_name,
                "item_author": it.item_author,
                "item_status": it.item_status,
                "starting_price": float(it.starting_price) if it.starting_price is not None else None,
                "reserve_price": float(it.reserve_price) if it.reserve_price is not None else None,
                "item_location": it.item_location,
                "item_box_code": it.item_box_code,
                "item_category": it.item_category,
                "item_notes": it.item_notes,
                "item_size": it.item_size,
                "item_image": it.item_image,
                "stockin_date": str(it.stockin_date) if it.stockin_date else None,
                "seller_code": it.seller_code,
                "item_material": it.item_material,
                "accessories": accessories
            })
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # [API] 在库编辑：PUT /api/items/<item_code>
    @app.route("/api/items/<item_code>", methods=["PUT"])
    def api_items_update(item_code):
        payload = request.json or {}
        session = get_session()
        try:
            it = session.get(Item, item_code)
            if not it:
                return jsonify({"error": "item 不存在"}), 404

            before = {
                "item_status": it.item_status,
                "starting_price": str(it.starting_price) if it.starting_price is not None else None,
                "reserve_price": str(it.reserve_price) if it.reserve_price is not None else None,
                "item_location": it.item_location,
                "item_box_code": it.item_box_code,
                "item_category": it.item_category,
                "item_notes": it.item_notes,
                "item_name": it.item_name,
                "item_author": it.item_author,
                "item_size": it.item_size,
                "item_image": it.item_image,
                "item_material": it.item_material,
            }

            # 金额/状态校验（万日元整数 & 起拍价 ≤ 底价）
            sp = it.starting_price
            rp = it.reserve_price

            if "starting_price" in payload:
                sp = to_int_or_none(payload.get("starting_price"))
            if "reserve_price" in payload:
                rp = to_int_or_none(payload.get("reserve_price"))

            # 两者同时有值时校验
            if sp is not None and rp is not None and sp > rp:
                raise ValueError("起拍价不能高于底价")

            it.starting_price = sp
            it.reserve_price = rp

            if "item_status" in payload:
                it.item_status = ensure_status(session, payload.get("item_status"))

            # 其他字段直写（不包含附属品）
            for k in ["item_location", "item_box_code", "item_category", "item_notes",
                      "item_name", "item_author", "item_size", "item_image", "item_material"]:
                if k in payload:
                    setattr(it, k, (payload.get(k) if payload.get(k) not in ("", None) else None))

            log_op(session, "item", item_code, "update", before=str(before), after=str(payload))
            session.commit()
            return jsonify({"ok": True})
        except ValueError as ve:
            session.rollback()
            return jsonify({"error": str(ve)}), 400
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # [API] 覆盖保存某件商品的附属品
    @app.route("/api/items/<item_code>/accessories", methods=["PUT"])
    def api_items_update_accessories(item_code):
        session = get_session()
        try:
            it = session.get(Item, item_code)
            if not it:
                return jsonify({"error": f"item 不存在: {item_code}"}), 404

            data = request.get_json(silent=True) or {}
            accessories = data.get("accessories", [])
            if isinstance(accessories, str):
                accessories = [s.strip() for s in accessories.replace(",", "、").split("、") if s.strip()]

            # 使用映射表覆盖保存
            session.execute(text("DELETE FROM item_accessories WHERE item_code = :c"), {"c": item_code})
            for name in accessories:
                session.execute(
                    text("INSERT INTO item_accessories(item_code, accessory_name) VALUES(:c, :n)"),
                    {"c": item_code, "n": name}
                )

            try:
                log_op(session, "item", item_code, "update_accessories", before=None, after=str(accessories))
            except Exception:
                pass

            session.commit()
            return jsonify({"ok": True})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # [API] 在库查询（分页+筛选+模糊）
    @app.route("/api/items", methods=["GET"], endpoint="api_items_index")
    def api_items_index():
        session = get_session()
        try:
            # 参数
            page = max(int(request.args.get("page", 1)), 1)
            page_size = min(max(int(request.args.get("page_size", 20)), 1), 100)
            # 兼容 keyword -> q
            q = (request.args.get("q") or request.args.get("keyword") or "").strip()
            seller_code = request.args.get("seller_code") or None
            status = request.args.get("status") or None
            status_group = request.args.get("status_group") or None  # 一级状态（在库 / 出库）
            category = request.args.get("category") or None
            # 新增筛选参数（列表页顶栏/表头筛选会用到）
            seller = request.args.get("seller") or None  # 出品人（姓名/代号模糊）
            seller_codes = request.args.get("seller_codes") or None  # 多选：逗号分隔
            stockin_date_from = request.args.get("stockin_date_from") or None
            stockin_date_to = request.args.get("stockin_date_to") or None
            box_code = request.args.get("box_code") or None
            loc = request.args.get("loc") or None
            item_name = request.args.get("item_name") or None
            # 兼容 size -> item_size
            item_size = request.args.get("item_size") or request.args.get("size") or None
            item_author = request.args.get("item_author") or None
            item_material = request.args.get("item_material") or None
            item_seal = request.args.get("item_seal") or None
            item_inscription = request.args.get("item_inscription") or None
            item_description = request.args.get("item_description") or None
            sp_min = request.args.get("sp_min") or None
            sp_max = request.args.get("sp_max") or None
            rp_min = request.args.get("rp_min") or None
            rp_max = request.args.get("rp_max") or None
            EMPTY = "__EMPTY__"

            query = session.query(Item)
            if seller_code:
                query = query.filter(Item.seller_code == seller_code)
            # 多个出品人（多选）：seller_codes=CODE1,CODE2,...
            # 多个出品人（多选）：seller_codes=CODE1,CODE2,...
            if seller_codes:
                arr = [s.strip().upper() for s in (seller_codes or '').split(',') if s.strip()]
                if arr:
                    query = query.filter(Item.seller_code.in_(arr))

            # 先按「一级状态」（在库 / 出库）筛选
            if status_group:
                from create_database import ItemStatus
                # 根据二级状态名，关联到 ItemStatus 拿到 group_name
                query = query.outerjoin(ItemStatus, Item.item_status == ItemStatus.item_status)
                query = query.filter(ItemStatus.group_name == status_group)

            # 再按「二级状态字符串」筛选
            if status:
                s = status.strip()
                if s == "EMPTY":
                    # item_status 为空或空字符串
                    query = query.filter((Item.item_status.is_(None)) | (Item.item_status == ""))
                elif s == "NON_EMPTY":
                    # item_status 非空
                    query = query.filter((Item.item_status.isnot(None)) & (Item.item_status != ""))
                else:
                    query = query.filter(Item.item_status == s)

            if category == EMPTY:
                query = query.filter((Item.item_category.is_(None)) | (Item.item_category == ""))
            elif category:
                query = query.filter(Item.item_category == category)


            from sqlalchemy import or_, and_
            from datetime import datetime

            if seller:  # 姓名/代号模糊匹配
                like = f"%{seller}%"
                query = query.filter(or_(Item.seller_name.ilike(like), Item.seller_code.ilike(like)))

            if stockin_date_from:
                try:
                    d = datetime.strptime(stockin_date_from, "%Y-%m-%d").date()
                    query = query.filter(Item.stockin_date >= d)
                except Exception:
                    pass

            if stockin_date_to:
                try:
                    d = datetime.strptime(stockin_date_to, "%Y-%m-%d").date()
                    query = query.filter(Item.stockin_date <= d)
                except Exception:
                    pass

            if box_code:
                like = f"%{box_code}%"
                query = query.filter((Item.item_box_code.ilike(like)))

            if loc == EMPTY:
                query = query.filter((Item.item_location.is_(None)) | (Item.item_location == ""))
            elif loc:
                like = f"%{loc}%"
                query = query.filter(Item.item_location.ilike(like))

            if item_name== EMPTY:
                query = query.filter((Item.item_name.is_(None)) | (Item.item_name == ""))
            elif item_name:
                like = f"%{item_name}%"
                query = query.filter(Item.item_name.ilike(like))


            if item_size== EMPTY:
                query = query.filter((Item.item_size.is_(None)) | (Item.item_size == ""))
            elif item_size:
                like = f"%{item_size}%"
                query = query.filter(Item.item_size.ilike(like))

            if item_author== EMPTY:
                query = query.filter((Item.item_author.is_(None)) | (Item.item_author == ""))
            elif item_author:
                like = f"%{item_author}%"
                query = query.filter(Item.item_author.ilike(like))

            if item_material== EMPTY:
                query = query.filter((Item.item_material.is_(None)) | (Item.item_material == ""))
            elif item_material:
                like = f"%{item_material}%"
                query = query.filter(Item.item_material.ilike(like))

            if item_seal== EMPTY:
                query = query.filter((Item.item_seal.is_(None)) | (Item.item_seal == ""))
            elif item_seal:
                like = f"%{item_seal}%"
                query = query.filter(Item.item_seal.ilike(like))

            if item_inscription== EMPTY:
                query = query.filter((Item.item_inscription.is_(None)) | (Item.item_inscription == ""))
            elif item_inscription:
                like = f"%{item_inscription}%"
                query = query.filter(Item.item_inscription.ilike(like))

            if item_description== EMPTY:
                query = query.filter((Item.item_description.is_(None)) | (Item.item_description == ""))
            elif item_description:
                like = f"%{item_description}%"
                query = query.filter(Item.item_description.ilike(like))
            # 价格范围（万日元，整数字符串）
            def _to_int(v):
                try:
                    return int(str(v))
                except Exception:
                    return None

            if request.args.get("sp_empty") == "1":
                query = query.filter(Item.starting_price.is_(None))
            if request.args.get("rp_empty") == "1":
                query = query.filter(Item.reserve_price.is_(None))

            _sp_min = _to_int(sp_min);
            _sp_max = _to_int(sp_max)
            _rp_min = _to_int(rp_min);
            _rp_max = _to_int(rp_max)
            if _sp_min is not None:
                query = query.filter(Item.starting_price >= _sp_min)
            if _sp_max is not None:
                query = query.filter(Item.starting_price <= _sp_max)
            if _rp_min is not None:
                query = query.filter(Item.reserve_price >= _rp_min)
            if _rp_max is not None:
                query = query.filter(Item.reserve_price <= _rp_max)

            if q:
                like = f"%{q}%"
                query = query.filter(
                    (Item.item_code.ilike(like)) |
                    (Item.item_name.ilike(like)) |
                    (Item.item_author.ilike(like)) |
                    (Item.item_description.ilike(like))
                )

            # 取出全部，做自然排序
            all_rows = query.all()

            def batch_key(x):
                # 保留你原有“按日期/出品人”的前两级
                prefix = (str(x.stockin_date or ""), code_to_number(x.seller_code or ""))
                return prefix + (item_code_nat_key_from_code(getattr(x, "item_code", None)),)

            all_rows.sort(key=batch_key)

            total = len(all_rows)
            start = (page - 1) * page_size
            end = start + page_size
            rows = all_rows[start:end]

            # ===== 批量读取出品人姓名映射（code -> name），来自 sellers 表 =====
            seller_codes = sorted({r.seller_code for r in rows if getattr(r, "seller_code", None)})
            seller_name_map = {}
            if seller_codes:
                pairs = (
                    session.query(Seller.seller_code, Seller.seller_name)
                    .filter(Seller.seller_code.in_(seller_codes))
                    .all()
                )
                for code, name in pairs:
                    seller_name_map[code] = name or code

            # ===== 批量查询附属品并映射到 {item_code: [name, ...]} =====
            from create_database import ItemAccessory  # 已有模型

            codes = [r.item_code for r in rows if getattr(r, "item_code", None)]
            acc_map = {}
            if codes:
                pairs = (
                    session.query(ItemAccessory.item_code, ItemAccessory.accessory_name)
                    .filter(ItemAccessory.item_code.in_(codes))
                    .order_by(ItemAccessory.item_code.asc(), ItemAccessory.accessory_name.asc())
                    .all()
                )
                for c, name in pairs:
                    acc_map.setdefault(c, []).append(name or "")

            def to_row(x):
                return {
                    "item_code": x.item_code,
                    "item_name": x.item_name,
                    "item_author": x.item_author,
                    "seller_code": x.seller_code,
                    "seller_name": seller_name_map.get(x.seller_code, x.seller_code),
                    "item_status": x.item_status,
                    "starting_price": float(x.starting_price) if x.starting_price is not None else None,
                    "reserve_price": float(x.reserve_price) if x.reserve_price is not None else None,
                    "item_location": x.item_location,
                    "item_box_code": x.item_box_code,
                    "item_category": x.item_category,
                    "item_image": x.item_image,
                    "stockin_date": str(x.stockin_date) if x.stockin_date else None,
                    "item_size": x.item_size,
                    "item_material": x.item_material,
                    "item_seal": x.item_seal,
                    "item_inscription": x.item_inscription,
                    "item_description": x.item_description,
                    "accessories_text": "、".join(acc_map.get(x.item_code, [])),  # 来自 item_accessories
                }

            return jsonify({
                "page": page, "page_size": page_size, "total": total,
                "items": [to_row(r) for r in rows]
            })

        finally:
            session.close()

    # [API] 新增单件
    @app.route("/api/items", methods=["POST"])
    def api_items_create():
        payload = request.json or {}
        required = ["item_code", "item_name", "seller_code", "stockin_date"]
        for r in required:
            if not payload.get(r):
                return jsonify({"error": f"缺少必填字段: {r}"}), 400

        session = get_session()
        try:
            if session.get(Item, payload["item_code"]):
                return jsonify({"error": "item_code 已存在"}), 400

            reserve_price = to_decimal_or_none(payload.get("reserve_price"))
            starting_price = to_decimal_or_none(payload.get("starting_price"))

            # 默认“在库”：若前端未传 item_status，则尝试设置为“在库”（需该状态存在于表 ItemStatus）
            status = None
            req_status = payload.get("item_status")
            if req_status:
                status = ensure_status(session, req_status)  # 校验并使用前端传入值
            else:
                try:
                    from create_database import ItemStatus
                    if session.get(ItemStatus, "待上拍"):
                        status = "待上拍"
                except Exception:
                    status = None  # 表里没有“在库”时，保持为 None，避免报错

            it = Item(
                item_code=payload["item_code"],
                item_accessories=payload.get("item_accessories"),
                item_name=payload["item_name"],
                item_size=payload.get("item_size"),
                item_image=payload.get("item_image"),
                is_in_box=bool(payload.get("is_in_box", False)),
                item_box_code=payload.get("item_box_code"),
                item_location=payload.get("item_location"),
                item_category=payload.get("item_category"),
                reserve_price=reserve_price,
                seller_name=payload.get("seller_name"),
                seller_code=payload["seller_code"],
                starting_price=starting_price,
                stockin_date=payload["stockin_date"],
                item_order=payload.get("item_order"),
                item_barcode=payload.get("item_barcode"),
                photo_date_shot=payload.get("photo_date_shot"),
                photo_date_detail=payload.get("photo_date_detail"),
                photo_date_ps=payload.get("photo_date_ps"),
                item_material=payload.get("item_material"),
                item_seal=payload.get("item_seal"),
                item_inscription=payload.get("item_inscription"),
                item_description=payload.get("item_description"),
                item_author=payload.get("item_author"),
                item_status=status,
                item_notes=payload.get("item_notes"),
            )
            session.add(it)
            from create_database import StockBatch
            sb = session.get(StockBatch, {"stockin_date": it.stockin_date, "seller_code": it.seller_code})
            if not sb:
                sb = StockBatch(stockin_date=it.stockin_date, seller_code=it.seller_code, stockin_count=0)
                session.add(sb)
            sb.stockin_count = (sb.stockin_count or 0) + 1
            log_op(session, "item", it.item_code, "create", after=str(payload))
            session.commit()
            return jsonify({"ok": True})
        except ValueError as ve:
            session.rollback()
            return jsonify({"error": str(ve)}), 400
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # === 新增：批量新建（支持“新增一件/多件”） ===
    @app.route("/api/items/bulk-create", methods=["POST"])
    def api_items_bulk_create():
        """
        body: { stockin_date: "YYYY-MM-DD", seller_code: "...", items: [{item_code: "..."}] }
        - 仅创建主键和批次外键，其它字段置空
        - 已存在的 item_code 跳过
        """
        payload = request.get_json(silent=True) or {}
        stockin_date = payload.get("stockin_date")
        seller_code = (payload.get("seller_code") or "").strip()
        items = payload.get("items") or []
        if not stockin_date or not seller_code or not isinstance(items, list) or not items:
            return jsonify({"error": "参数不完整"}), 400

        from datetime import datetime
        try:
            stockin_date_obj = datetime.strptime(stockin_date, "%Y-%m-%d").date()
        except Exception:
            return jsonify({"error": "stockin_date 必须为 YYYY-MM-DD"}), 400

        session = get_session()
        try:
            created = 0
            # 计算默认状态（批量复用）
            default_status = None
            try:
                from create_database import ItemStatus
                if session.get(ItemStatus, "待上拍"):
                    default_status = "待上拍"
            except Exception:
                default_status = None
            for row in items:
                code = (row.get("item_code") or "").strip()
                if not code:
                    continue
                if session.get(Item, code):
                    continue
                it = Item(
                    item_code=code,
                    stockin_date=stockin_date_obj,
                    seller_code=seller_code,
                    item_name=None, item_size=None, item_location=None,
                    item_category=None, item_image=None, item_notes=None,
                    starting_price=None, reserve_price=None, item_status=default_status,
                    item_box_code=None, item_author=None
                )
                session.add(it)
                created += 1
                try:
                    log_op(session, "item", code, "create", before=None, after=None)
                except Exception:
                    pass

            from create_database import StockBatch
            sb = session.get(StockBatch, {"stockin_date": stockin_date_obj, "seller_code": seller_code})
            if not sb:
                sb = StockBatch(stockin_date=stockin_date_obj, seller_code=seller_code, stockin_count=0)
                session.add(sb)
            sb.stockin_count = (sb.stockin_count or 0) + created

            session.commit()
            return jsonify({"ok": True, "created": created})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # === 新增：删除单件（前端行内“删除”按钮用） ===
    @app.route("/api/items/<item_code>", methods=["DELETE"])
    def api_items_delete(item_code):
        session = get_session()
        try:
            it = session.get(Item, item_code)
            if not it:
                # 幂等：视为已删除
                return jsonify({"ok": True, "note": "already deleted"}), 200

            # 1) 清理附属品映射表
            session.execute(text("DELETE FROM item_accessories WHERE item_code = :c"), {"c": item_code})

            # 2) 可选：删除该物品图片目录（若存在）
            try:
                subdir = os.path.join(UPLOAD_ROOT, item_code)
                if os.path.isdir(subdir):
                    shutil.rmtree(subdir)
            except Exception:
                # 删除文件失败不影响主事务
                pass

            from create_database import StockBatch
            sb = session.get(StockBatch, {"stockin_date": it.stockin_date, "seller_code": it.seller_code})
            if sb and sb.stockin_count is not None:
                sb.stockin_count = max(0, int(sb.stockin_count) - 1)

            # 3) 删除物品本体
            session.delete(it)
            try:
                log_op(session, "item", item_code, "delete", before=None, after=None)
            except Exception:
                pass

            session.commit()
            return jsonify({"ok": True})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # === 批次内“删除无信息物品”：
    # 规则：除 item_code/stockin_date/seller_code/item_status 外，其余字段全部为空，
    #      且无任何附属品(item_accessories) 的条目才允许删除
    @app.route("/api/items/cleanup-empty", methods=["POST"])
    def api_items_cleanup_empty():
        payload = request.get_json(silent=True) or {}
        stockin_date = (payload.get("stockin_date") or "").strip()
        seller_code = (payload.get("seller_code") or "").strip()
        dry_run = bool(payload.get("dry_run", False))

        if not stockin_date or not seller_code:
            return jsonify({"error": "缺少 stockin_date 或 seller_code"}), 400

        session = get_session()
        try:
            # 读取该批次的必要字段
            rows = session.execute(text("""
                SELECT i.item_code,
                       i.item_name, i.item_size, i.item_location, i.item_box_code,
                       i.item_category, i.item_image, i.item_notes,
                       i.starting_price, i.reserve_price, i.item_author
                FROM items i
                WHERE i.stockin_date = :d AND i.seller_code = :s
            """), {"d": stockin_date, "s": seller_code}).mappings().all()

            def is_blank(v):
                return v is None or (isinstance(v, str) and v.strip() == "")

            candidates = []
            for r in rows:
                code = r["item_code"]
                # 有任一“信息字段”不为空，则不可删
                info_fields = ["item_name", "item_size", "item_location", "item_box_code",
                               "item_category", "item_image", "item_notes",
                               "starting_price", "reserve_price", "item_author"]
                has_info = any(not is_blank(r[k]) for k in info_fields)

                if has_info:
                    continue

                # 有附属品记录也不可删
                acc_cnt = session.execute(
                    text("SELECT COUNT(*) FROM item_accessories WHERE item_code = :c"),
                    {"c": code}
                ).scalar() or 0
                if int(acc_cnt) > 0:
                    continue

                candidates.append(code)

            if dry_run:
                return jsonify({"ok": True, "codes": candidates, "total": len(candidates)})

            # 执行删除（与单件删除保持一致：先删映射表、再删主表、批次件数回退）
            deleted = 0
            from create_database import StockBatch
            for code in candidates:
                it = session.get(Item, code)
                if not it:
                    continue
                # 附属品映射
                session.execute(text("DELETE FROM item_accessories WHERE item_code = :c"), {"c": code})
                # 批次件数回退
                sb = session.get(StockBatch, {"stockin_date": it.stockin_date, "seller_code": it.seller_code})
                if sb and sb.stockin_count is not None:
                    try:
                        sb.stockin_count = max(0, int(sb.stockin_count) - 1)
                    except Exception:
                        pass
                # 删除物品
                session.delete(it)
                try:
                    log_op(session, "item", code, "delete_empty", before=None, after=None)
                except Exception:
                    pass
                deleted += 1

            session.commit()
            return jsonify({"ok": True, "deleted": deleted, "codes": candidates})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # ========================= 导出/预检通用工具 =========================
    def _fetch_batch_items(session, stockin_date: str, seller_code: str):
        """读取指定批次的简要信息（导出用）"""
        rows = session.execute(text("""
            SELECT i.item_code, i.item_name, i.starting_price, i.reserve_price,
                   i.item_notes, i.item_image, i.seller_code, i.stockin_date
            FROM items i
            WHERE i.stockin_date = :d AND i.seller_code = :s
            ORDER BY i.item_code
        """), {"d": stockin_date, "s": seller_code}).mappings().all()
        return [dict(r) for r in rows]

    def _fetch_seller_name(session, seller_code: str) -> str:
        s = session.get(Seller, seller_code)
        return getattr(s, "seller_name", "") if s else ""

    def _format_batch_code(stockin_date: str, seller_code: str) -> str:
        """YYMMDD_SCODE 形式，如 250324_C"""
        dt = datetime.strptime(stockin_date, "%Y-%m-%d")
        return dt.strftime("%y%m%d") + f"_{seller_code}"

    # ====== 简繁体转换 & 拼音首字母（可选依赖，未安装则降级为原名匹配） ======
    try:
        # pip install opencc-python-reimplemented
        from opencc import OpenCC
        _cc_t2s = OpenCC('t2s')  # 繁转简
        _cc_s2t = OpenCC('s2t')  # 简转繁
    except Exception:
        _cc_t2s = _cc_s2t = None

    def to_simplified(s: str) -> str:
        if not s: return s
        if _cc_t2s:
            try:
                return _cc_t2s.convert(s)
            except Exception:
                pass
        return s

    def to_traditional(s: str) -> str:
        if not s: return s
        if _cc_s2t:
            try:
                return _cc_s2t.convert(s)
            except Exception:
                pass
        return s

    try:
        # pip install pypinyin
        from pypinyin import lazy_pinyin, Style
    except Exception:
        lazy_pinyin = None
        Style = None

    def pinyin_initials(s: str) -> str:
        """
        取每个汉字的拼音首字母并大写，如“洪世国” -> HSG。
        非汉字忽略；无 pypinyin 时返回空串（前端仍可用中文匹配）。
        """
        if not s or not lazy_pinyin or not Style:
            return ""
        initials = []
        for py in lazy_pinyin(s, style=Style.NORMAL, errors='ignore'):
            if py:
                initials.append(py[0])
        return "".join(initials).upper()

    def _check_missing_images(items):
        """返回缺图的 item_code 列表（无路径或文件不存在）"""
        missing = []
        for it in items:
            rel = (it.get("item_image") or "").strip()
            if not rel:
                missing.append(it["item_code"]);
                continue
            fs = _abs_path_from_web(rel)
            if not os.path.exists(fs):
                missing.append(it["item_code"])
        return missing

    def _export_excel(stockin_date: str, seller_code: str, seller_name: str, items):
        # 依赖检查
        if Workbook is None:
            raise RuntimeError("缺少依赖：openpyxl；请安装：pip install openpyxl pillow")

        # ------------ 工具 & 导入 ------------
        import io, os, re
        from datetime import datetime
        from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
        from openpyxl.worksheet.page import PageMargins

        # 从 config 读取 LOGO 路径（可相对 / 绝对）
        try:
            from config import LOGO_PATH as _CFG_LOGO
        except Exception:
            _CFG_LOGO = None

        # ===== 这里一个参数控制“图片单元格”的边长（像素）=====
        IMG_CELL_PX = 100  # ← 想要正方形 110px/120px，只改这里

        # px ↔ pt / 列宽换算
        def px_to_pt(px):  # 96 dpi → 72 pt/inch
            return px * 72 / 96.0

        def px_to_colwidth(px):  # pixels ≈ 7*width + 5
            return max(1, round((px - 5) / 7.0, 2))

        def colwidth_to_px(w):  # width → pixels
            return int(round(7 * float(w) + 5))

        # 自然排序（确保 1,2,3…）
        def nat_key(code: str):
            s = str(code or "")
            m = re.search(r"^(.*?)[_\-]?(\d+)$", s)
            if m: return (m.group(1), int(m.group(2)))
            return (s, 0)

        items_sorted = sorted(items, key=lambda it: nat_key(it.get("item_code")))

        # ------------ 新建工作簿 ------------
        wb = Workbook()
        ws = wb.active
        ws.title = "批次清单"

        # ------------ 页面设置（A4 + 边距 + 居中 + 头/脚）------------
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(left=0.3937, right=0.3937, top=0.6, bottom=0.6, header=0.3, footer=0.3)
        ws.print_options.horizontalCentered = True

        # 顶端重复行（打印标题行）：第 2~5 行
        try:
            ws.print_title_rows = '2:5'
        except Exception:
            pass

        # ------------ 页眉/页脚（与正文同字体）------------
        font_name = "微软雅黑"
        batch_code = _format_batch_code(stockin_date, seller_code)
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M')

        hdr_left = f'&"{font_name},Regular"&10生成时间：{now_str}'
        hdr_right = f'&"{font_name},Regular"&10批次：{batch_code}'

        ftr_c = f'&"{font_name},Regular"&10第&P页 / 共&N页'

        # LOGO：页眉中间（如路径存在）
        center_header_code = "&G"
        logo_added = False
        logo_fs = None
        if _CFG_LOGO:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            logo_fs = _CFG_LOGO if os.path.isabs(_CFG_LOGO) else os.path.join(base_dir, _CFG_LOGO)
            if not os.path.exists(logo_fs):
                logo_fs = None

        try:
            ws.header_footer.left_header = hdr_left
            ws.header_footer.right_header = hdr_right
            ws.header_footer.center_footer = ftr_c
            if logo_fs:
                try:
                    logo_img = XLImage(logo_fs)
                    target_h = 36  # 页眉目标高度（px）
                    if Image:
                        with Image.open(logo_fs) as _pil:
                            ow, oh = _pil.size
                            scale = min(1.0, target_h / max(1, oh))
                            logo_img.height = int(oh * scale)
                            logo_img.width = int(ow * scale)
                    ws.header_footer.add_image(logo_img, 'C')
                    ws.header_footer.center_header = center_header_code
                    logo_added = True
                except Exception:
                    pass
            if not logo_added:
                ws.header_footer.center_header = ""
        except Exception:
            try:
                ws.oddHeader.left.text = hdr_left
                ws.oddHeader.right.text = hdr_right
                ws.oddFooter.center.text = ftr_c
            except Exception:
                pass

        # ------------ 字体/样式 ------------
        title_font = Font(name=font_name, size=16, bold=True)
        info_font = Font(name=font_name, size=11)
        th_font = Font(name=font_name, size=11, bold=True)
        td_font = Font(name=font_name, size=10)

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="000000")
        border = Border(top=thin, bottom=thin, left=thin, right=thin)

        # ------------ 列顺序/宽度 ------------
        COLS = ["内部编号", "图片", "名称", "起拍价", "底价", "备注"]
        ws.append(COLS)  # 先占位一行
        header_row = 1

        # 图片列宽用 IMG_CELL_PX 统一控制（正方形单元格）
        width_A = 14.0  # 编号
        width_B = px_to_colwidth(IMG_CELL_PX)  # 图片（≈IMG_CELL_PX 像素）
        width_F = 15.0  # 备注
        width_D = round(width_F * 0.8, 1)  # 起拍价
        width_E = width_D  # 底价
        width_C = 24.0  # 名称

        ws.column_dimensions['A'].width = width_A
        ws.column_dimensions['B'].width = width_B
        ws.column_dimensions['C'].width = width_C
        ws.column_dimensions['D'].width = width_D
        ws.column_dimensions['E'].width = width_E
        ws.column_dimensions['F'].width = width_F

        # ------------ 标题/信息区（行1~3）------------
        ws.insert_rows(1, amount=3)
        ws.merge_cells("A1:F1")
        ws["A1"] = "吉祥オークション"
        ws["A1"].font = title_font
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = px_to_pt(36)

        ws.merge_cells("A2:C2")
        ws.merge_cells("E2:F2")
        display_name = (seller_name or seller_code).strip()
        ws["A2"] = f"     {display_name} 様"  # 5 个空格
        ws["E2"] = f"总件数：{len(items_sorted)}"
        ws["A2"].font = info_font;
        ws["A2"].alignment = left
        ws["E2"].font = info_font;
        ws["E2"].alignment = center
        ws.row_dimensions[2].height = px_to_pt(26)
        ws.row_dimensions[3].height = px_to_pt(8)

        # ------------ 表头（第4行）------------
        header_row = 4
        for j, h in enumerate(COLS, start=1):
            cell = ws.cell(row=header_row, column=j, value=h)
            cell.font = th_font
            cell.alignment = center
            cell.border = border
            cell.fill = PatternFill("solid", fgColor="F5F5F5")
        ws.row_dimensions[header_row].height = px_to_pt(28)

        # ---- 单位行（第5行）：A/B/C/F 纵向合并；D/E 放“（万日元）”；D4↔D5 & E4↔E5 无分隔线 ----
        sub_row = header_row + 1
        ws.row_dimensions[sub_row].height = px_to_pt(18)

        # A/B/C/F：把第4~5行纵向合并（视觉上就是“同一格里换行”）
        for col_letter in ('A', 'B', 'C', 'F'):
            ws.merge_cells(f"{col_letter}{header_row}:{col_letter}{sub_row}")

        # D/E：小字单位，背景与表头一致；去掉 D4→D5 与 E4→E5 的横线
        for col in (4, 5):
            top = ws.cell(row=header_row, column=col)  # D4/E4
            bot = ws.cell(row=sub_row, column=col, value="（万日元）")  # D5/E5

            # 上面这一格保留左右+上边框，去掉底边（与下面一格“无缝”）
            top.border = Border(left=thin, right=thin, top=thin, bottom=Side(style=None))
            top.alignment = center

            # 下面这一格：小一号，居中；去掉上边框；底边框保留
            bot.font = Font(name=font_name, size=9)
            bot.alignment = center
            bot.fill = PatternFill("solid", fgColor="F5F5F5")
            bot.border = Border(left=thin, right=thin, top=Side(style=None), bottom=thin)

        # ------------ 数据（第5行起）------------
        row_idx = header_row + 2
        MIN_PAD = 2  # 最小边距（px）

        for it in items_sorted:
            ws.cell(row=row_idx, column=1, value=it.get("item_code") or "")
            ws.cell(row=row_idx, column=2, value=None)
            ws.cell(row=row_idx, column=3, value=it.get("item_name") or "")
            sp = it.get("starting_price")
            ws.cell(row=row_idx, column=4, value=(float(sp) if sp not in (None, "") else None))
            rp = it.get("reserve_price")
            ws.cell(row=row_idx, column=5, value=(float(rp) if rp not in (None, "") else None))
            ws.cell(row=row_idx, column=6, value=it.get("item_notes") or "")

            # —— 强制该行行高 = IMG_CELL_PX 像素（保证正方形）——
            ws.row_dimensions[row_idx].height = px_to_pt(IMG_CELL_PX)
            # 单元格实际像素：行高固定=IMG_CELL_PX；列宽按设定换算得到像素
            b_width_chars = ws.column_dimensions['B'].width or width_B
            cell_w_px = colwidth_to_px(b_width_chars)
            cell_h_px = IMG_CELL_PX

            # 可用盒子（至少留 MIN_PAD 边距）
            avail_w = max(1, cell_w_px - 2 * MIN_PAD)
            avail_h = max(1, cell_h_px - 2 * MIN_PAD)

            # ===== 图片：等比缩放 contain + 单元格内居中，不改变源文件像素 =====
            rel = it.get("item_image")
            placed = False

            try:
                if rel and XLImage is not None:
                    fs = _abs_path_from_web(rel) if callable(globals().get("_abs_path_from_web")) else rel
                    if os.path.exists(fs):
                        # 读原始字节
                        with open(fs, "rb") as _f:
                            raw = _f.read()
                        # EXIF 自动转正（不改变像素分辨率）
                        try:
                            from PIL import Image as PILImage, ImageOps
                            bio = io.BytesIO(raw)
                            with PILImage.open(bio) as im:
                                im = ImageOps.exif_transpose(im)
                                out = io.BytesIO()
                                fmt = im.format or "PNG"
                                im.save(out, format=fmt, quality=95)
                                raw = out.getvalue()
                        except Exception:
                            pass
                        # 构造图片对象，拿到原始像素尺寸
                        buf = io.BytesIO(raw)
                        ximg = XLImage(buf)
                        img_w_px = int(ximg.width)
                        img_h_px = int(ximg.height)
                        # 目标：本项目图片列固定是 B 列
                        target_col_letter = 'B'

                        # 取得 B列、当前行 的单元格像素盒子大小（列宽×行高，单位：px）
                        cell_w_px, cell_h_px = get_cell_box_px(ws, target_col_letter, row_idx)
                        # —— 等比例缩放（contain）——
                        scale_w = cell_w_px / max(1, img_w_px)
                        scale_h = cell_h_px / max(1, img_h_px)
                        scale = min(scale_w, scale_h)
                        if not ALLOW_UPSCALE:
                            scale = min(scale, 1.0)  # 只缩小不放大，保证清晰
                        disp_w_px = int(round(img_w_px * scale))
                        disp_h_px = int(round(img_h_px * scale))

                        # —— 居中偏移（以单元格左上角为原点）——
                        off_x_px = max(0, (cell_w_px - disp_w_px) // 2)
                        off_y_px = max(0, (cell_h_px - disp_h_px) // 2)
                        # —— OneCellAnchor 锚在 B列 的该行 ——（0-based：B=1, 行=row_idx-1）
                        col0 = 1
                        row0 = row_idx - 1

                        marker = AnchorMarker(
                            col=col0,
                            colOff=off_x_px * EMU_PER_PX,
                            row=row0,
                            rowOff=off_y_px * EMU_PER_PX
                        )
                        # 关键：ext 必须是 XDRPositiveSize2D（不是 geometry.PositiveSize2D）
                        ext = XDRPositiveSize2D(disp_w_px * EMU_PER_PX, disp_h_px * EMU_PER_PX)
                        ximg.anchor = OneCellAnchor(_from=marker, ext=ext)
                        ws.add_image(ximg)
                        placed = True
            except Exception as e:
                print("图片放置失败:", e)

            if not placed:
                ws.cell(row=row_idx, column=2, value="(无图片或加载失败)").alignment = center

            # 对齐/边框/数字格式
            for col in range(1, 7):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = td_font
                cell.border = border
                if col in (1, 4, 5):
                    cell.alignment = center
                elif col in (3, 6):
                    cell.alignment = left
            ws.cell(row=row_idx, column=4).number_format = '#,##0'
            ws.cell(row=row_idx, column=5).number_format = '#,##0'

            row_idx += 1

        # 冻结窗格（保留到表头）
        ws.freeze_panes = "A6"

        # ------------ 输出 ------------
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        fname = f"{_format_batch_code(stockin_date, seller_code)}.xlsx"
        return bio, fname

    def _export_pdf_via_excel_to_pdf(stockin_date: str, seller_code: str, seller_name: str, items):
        """
        新版 PDF 导出流程：
        1) 复用 _export_excel 生成 Excel（二进制内存流）；
        2) 写入系统临时目录为 .xlsx 临时文件；
        3) 使用 excel_to_pdf.convert_excel_to_pdf 直接转换成 PDF，并保存到项目 exports/pdf/ 目录；
        4) 将已保存的 PDF 再读出并返回给浏览器下载。

        保存路径：{BASE_DIR}/exports/pdf/{stockin_date}_{seller_code}.pdf
        """
        import io, os

        # 1) 先生成 Excel 的二进制流（仍然使用你现有的 _export_excel）
        bio_xlsx, xlsx_fname = _export_excel(stockin_date, seller_code, seller_name, items)
        bio_xlsx.seek(0)

        # 2) 写入系统临时目录，得到一个临时 .xlsx 路径（供 Excel COM 转换使用）
        tmp_dir = Path(tempfile.gettempdir())
        tmp_xlsx = tmp_dir / f"{stockin_date}_{seller_code}.xlsx"
        with open(tmp_xlsx, "wb") as f:
            f.write(bio_xlsx.read())

        # 3) 目标 PDF 永久保存目录：{BASE_DIR}/exports/pdf/
        export_dir = Path(BASE_DIR) / "exports" / "pdf"
        export_dir.mkdir(parents=True, exist_ok=True)
        out_pdf = export_dir / f"{stockin_date}_{seller_code}.pdf"

        # 4) 调用你提供的转换函数进行转换与保存
        #    注意：convert_excel_to_pdf 内部会严格按 Excel 打印区域导出 PDF（见 excel_to_pdf.py 文档）
        convert_excel_to_pdf(tmp_xlsx, out_pdf, open_visible=False)  # 生成并保存 PDF  →  项目目录

        # 5) 清理临时 xlsx（尽量不留垃圾文件）
        try:
            tmp_xlsx.unlink(missing_ok=True)
        except Exception:
            pass

        # 6) 供浏览器下载：把已保存的 PDF 读回内存
        bio_pdf = io.BytesIO(out_pdf.read_bytes())
        bio_pdf.seek(0)

        # 下载文件名：与旧逻辑一致（或保持清晰）
        download_name = out_pdf.name
        return bio_pdf, download_name

    # ====== 新增：中文字体注册（一次性） ======
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    def _get_cjk_font():
        """
        返回一个可显示中文的字体名；优先尝试系统/常见 CJK 字体；
        若不可用，回退到内置 CID 字体 STSong-Light（无需外部字体文件）。
        """
        # 缓存到全局避免重复注册
        global _CJK_FONT_NAME
        try:
            _CJK_FONT_NAME  # noqa
        except NameError:
            _CJK_FONT_NAME = None

        if _CJK_FONT_NAME:
            return _CJK_FONT_NAME

        # 1) 优先尝试常见的可用字体（如已部署了 Noto / 思源）
        candidates = [
            # (字体内部名, 绝对路径或 None 仅按名尝试)
            ("NotoSansCJKsc-Regular", "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Regular.otf"),
            ("SourceHanSansCN-Regular", "/usr/share/fonts/opentype/adobe-source-han-sans/SourceHanSansCN-Regular.otf"),
            ("SourceHanSansCN-Regular", "/usr/share/fonts/opentype/source-han-sans/SourceHanSansCN-Regular.otf"),
            ("PingFang SC", None),  # macOS
            ("Microsoft YaHei", None),  # Windows
            ("SimSun", None),  # Windows
            ("WenQuanYi Zen Hei", None)
        ]
        for name, path in candidates:
            try:
                if path and os.path.exists(path):
                    pdfmetrics.registerFont(TTFont(name, path))
                    _CJK_FONT_NAME = name
                    return _CJK_FONT_NAME
                # 仅按字体名尝试（如果系统字体可被 freetype 找到）
                pdfmetrics.registerFont(TTFont(name, name))
                _CJK_FONT_NAME = name
                return _CJK_FONT_NAME
            except Exception:
                pass

        # 2) 回退到内置 CID 字体 —— 支持中文、无需字体文件
        try:
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
            _CJK_FONT_NAME = 'STSong-Light'
            return _CJK_FONT_NAME
        except Exception:
            # 兜底（不会显示中文，但至少不报错）
            _CJK_FONT_NAME = 'Helvetica'
            return _CJK_FONT_NAME

    # ====== 新增：带“第 X / Y 页”的 Canvas ======
    from reportlab.pdfgen import canvas as _rl_canvas

    class NumberedCanvas(_rl_canvas.Canvas):
        """在保存时为每一页画上“第 x / y 页”的页码"""

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            page_count = len(self._saved_page_states) + 1  # 最后一页未调用 showPage 时也要计数
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self._draw_page_number(page_count)
                super().showPage()
            # 最后一页
            self._draw_page_number(page_count)
            super().save()

        def _draw_page_number(self, page_count):
            font_name = _get_cjk_font()
            self.setFont(font_name, 10)
            self.setFillColorRGB(0.35, 0.35, 0.35)
            # 页脚右下角：第 X / Y 页
            w, h = self._pagesize
            self.drawRightString(w - 36, 24, f"第 {self._pageNumber} / {page_count} 页")


    # ============================================================================
    # 六、Stock Batches（批次）相关接口
    # ============================================================================

    # [API] 批次列表
    @app.route("/api/stock-batches", methods=["GET"])
    def api_stock_batches():
        from create_database import StockBatch, Item, Seller, ItemStatus
        session = get_session()
        try:
            batches = session.query(StockBatch).all()

            result = []
            for b in batches:
                date_key = b.stockin_date
                scode = b.seller_code

                total_q = session.query(Item).filter(Item.stockin_date == date_key,
                                                     Item.seller_code == scode)
                total_items = total_q.count()
                instock_items = (
                    total_q
                    .outerjoin(ItemStatus, Item.item_status == ItemStatus.item_status)
                    .filter(ItemStatus.group_name == '在库')
                    .count()
                )

                sname = None
                s = session.get(Seller, scode)
                if s:
                    sname = s.seller_name

                result.append({
                    "stockin_date": str(date_key),
                    "seller_code": scode,
                    "seller_name": sname,
                    "stockin_count": b.stockin_count,
                    "total_items": total_items,
                    "in_stock": instock_items,
                    "has_physical_list": bool(b.has_physical_list),
                    "stockin_receiver": b.stockin_receiver,
                    "stockin_staff": b.stockin_staff,
                })

            # 日期倒序 + 出品人字母序
            result.sort(key=lambda x: (x["stockin_date"], x["seller_code"]), reverse=True)
            return jsonify(result)
        finally:
            session.close()

    # [API] 批次导出前的“缺图”预检
    @app.route("/api/batches/<stockin_date>/<seller_code>/precheck", methods=["GET"])
    def api_batch_precheck(stockin_date, seller_code):
        session = get_session()
        try:
            items = _fetch_batch_items(session, stockin_date, seller_code)
            missing = _check_missing_images(items)
            return jsonify({"ok": True, "total": len(items), "missing": missing})
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # [下载] 导出 Excel / PDF
    @app.route("/download/batches/<stockin_date>/<seller_code>.<ext>", methods=["GET"])
    def download_batch(stockin_date, seller_code, ext):
        session = get_session()
        try:
            items = _fetch_batch_items(session, stockin_date, seller_code)
            sname = _fetch_seller_name(session, seller_code)
            missing = _check_missing_images(items)
            if missing:
                # 直接返回 400，前端在点击前会先预检，这里是双保险
                return jsonify({"ok": False, "error": "图片缺失", "missing": missing}), 400

            if ext.lower() == "xlsx":
                bio, fname = _export_excel(stockin_date, seller_code, sname, items)
                return send_file(bio, as_attachment=True, download_name=fname,
                                 mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            elif ext.lower() == "pdf":
                bio, fname = _export_pdf_via_excel_to_pdf(stockin_date, seller_code, sname, items)
                return send_file(bio, as_attachment=True, download_name=fname, mimetype="application/pdf")
            else:
                return jsonify({"error": "不支持的格式"}), 400
        except RuntimeError as re:
            # 缺依赖的友好提示
            return jsonify({"ok": False, "error": str(re)}), 500
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # [API] 新建批次
    @app.route("/api/stock-batches", methods=["POST"])
    def api_stock_batches_create():
        from create_database import StockBatch
        payload = request.json or {}
        required = ["stockin_date", "seller_code"]
        for r in required:
            if not payload.get(r):
                return jsonify({"error": f"缺少必填字段: {r}"}), 400
        session = get_session()
        try:
            exist = session.get(StockBatch,
                                {"stockin_date": payload["stockin_date"], "seller_code": payload["seller_code"]})
            if exist:
                return jsonify({"ok": True, "message": "批次已存在"}), 200

            sb = StockBatch(
                stockin_date=payload["stockin_date"],
                seller_code=payload["seller_code"],
                stockin_count=payload.get("stockin_count") or None,
                has_physical_list=bool(payload.get("has_physical_list", False)),
                stockin_receiver=payload.get("stockin_receiver"),
                stockin_staff=payload.get("stockin_staff"),
            )
            session.add(sb)
            log_op(session, "stock_batch", f'{payload["stockin_date"]}|{payload["seller_code"]}', "create",
                   after=str(payload))
            session.commit()
            return jsonify({"ok": True})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # ============================================================================
    # 七、Sellers（出品人）相关接口
    # ============================================================================

    # [API] 迷你出品人下拉（增强：返回 keys 用于简繁体与拼音首字母匹配）
    @app.route("/api/sellers/mini")
    def api_sellers_mini():
        session = get_session()
        try:
            rows = session.query(Seller).all()
            rows.sort(key=lambda s: code_to_number(s.seller_code or ""))

            result = []
            for s in rows:
                code = (s.seller_code or "").upper()
                name = (s.seller_name or "").strip()

                simp = to_simplified(name)  # 简体
                trad = to_traditional(name)  # 繁体
                abbr = pinyin_initials(name)  # 拼音缩写（如 HSG）

                keys = set()
                for k in (name, simp, trad, abbr):
                    k = (k or "").strip()
                    if not k:
                        continue
                    keys.add(k)  # 原样
                    keys.add(k.upper())  # 大写，便于 HSG/hsg 都匹配

                result.append({
                    "seller_code": code,
                    "seller_name": name,
                    "keys": sorted(keys)  # 前端用于模糊匹配
                })

            return jsonify(result)
        finally:
            session.close()

    # [API] 出品人统计：每个出品人对应的在库物品数量
    @app.route("/api/sellers/stats", methods=["GET"])
    def api_sellers_stats():
        session = get_session()
        try:
            from sqlalchemy import text
            # 只统计「在库」分组的物品数量
            sql = text("""
                       SELECT s.seller_code, s.seller_name, COALESCE(cnt.c, 0) AS item_count
                       FROM sellers s
                                LEFT JOIN (SELECT i.seller_code, COUNT(*) AS c
                                           FROM items i
                                                    LEFT JOIN item_statuses st ON st.item_status = i.item_status
                                           WHERE st.group_name = '在库'
                                           GROUP BY i.seller_code) cnt ON cnt.seller_code = s.seller_code
                       ORDER BY s.seller_code ASC
                       """)
            rows = session.execute(sql).fetchall()
            data = [
                {
                    "seller_code": r[0],
                    "seller_name": r[1],
                    "item_count": int(r[2] or 0)
                } for r in rows
            ]
            return jsonify(data)
        finally:
            session.close()

    # [API] 诊断：检查简繁体/拼音依赖，并返回当前 Python 解释器与样例转换
    @app.route("/api/_diag/deps")
    def api_diag_deps():
        info = {"opencc_available": False, "pypinyin_available": False, "details": {}}

        # opencc 测试（有则做一次样例转换）
        try:
            from opencc import OpenCC
            cc1, cc2 = OpenCC('t2s'), OpenCC('s2t')
            info["opencc_available"] = True
            info["details"]["t2s_國"] = cc1.convert("國")  # 期望 "国"
            info["details"]["s2t_国"] = cc2.convert("国")  # 期望 "國"
        except Exception as e:
            info["details"]["opencc_error"] = str(e)

        # pypinyin 测试（有则给出“洪世國”的缩写）
        try:
            from pypinyin import lazy_pinyin, Style
            abb = ''.join(py[0] for py in lazy_pinyin("洪世國", style=Style.NORMAL, errors='ignore')).upper()
            info["pypinyin_available"] = True
            info["details"]["pinyin_HSG"] = abb  # 期望 "HSG"
        except Exception as e:
            info["details"]["pinyin_error"] = str(e)

        # 当前解释器路径
        import sys
        info["python_executable"] = sys.executable
        return jsonify(info)

    # [API] 出品人列表
    @app.route("/api/sellers", methods=["GET"])
    def api_sellers():
        session = get_session()
        try:
            sellers = session.query(Seller).all()
            sellers.sort(key=lambda s: code_to_number(s.seller_code or ""))
            result = []
            for s in sellers:
                result.append({
                    "seller_code": s.seller_code,
                    "seller_name": s.seller_name,
                    "seller_percent": float(s.seller_percent) if s.seller_percent is not None else None,
                    "seller_penalty_ratio": float(
                        s.seller_penalty_ratio) if s.seller_penalty_ratio is not None else None,
                    "is_catalog_fee_required": bool(s.is_catalog_fee_required),
                    "is_tax_deductible": bool(s.is_tax_deductible),
                    "seller_tax_code": s.seller_tax_code,
                    "seller_payment_account": s.seller_payment_account,
                    "seller_phone": s.seller_phone,
                    "seller_address": s.seller_address,
                    "seller_notes": s.seller_notes,
                })
            return jsonify(result)
        finally:
            session.close()

    # [API] 获取单个出品人
    @app.route("/api/sellers/<seller_code>", methods=["GET"])
    def api_sellers_get_one(seller_code):
        session = get_session()
        try:
            s = session.get(Seller, seller_code)
            if not s:
                return jsonify({"error": "seller 不存在"}), 404
            data = {
                "seller_code": s.seller_code,
                "seller_name": s.seller_name,
                "seller_percent": float(s.seller_percent) if s.seller_percent is not None else None,
                "seller_penalty_ratio": float(s.seller_penalty_ratio) if s.seller_penalty_ratio is not None else None,
                "is_catalog_fee_required": bool(s.is_catalog_fee_required),
                "is_tax_deductible": bool(s.is_tax_deductible),
                "seller_tax_code": s.seller_tax_code,
                "seller_payment_account": s.seller_payment_account,
                "seller_phone": s.seller_phone,
                "seller_address": s.seller_address,
                "seller_notes": s.seller_notes,
            }
            return jsonify(data)
        finally:
            session.close()

    # [API] 新建出品人
    @app.route("/api/sellers", methods=["POST"])
    def api_sellers_create():
        payload = request.json or {}
        required = ["seller_code", "seller_name"]
        for r in required:
            if not payload.get(r):
                return jsonify({"error": f"缺少必填字段: {r}"}), 400
        session = get_session()
        try:
            # 重号/重名校验
            if session.get(Seller, payload["seller_code"]):
                return jsonify({"error": "seller_code 已存在"}), 400
            exists_name = session.query(Seller).filter(Seller.seller_name == payload["seller_name"]).first()
            if exists_name:
                return jsonify({"error": "seller_name 已存在"}), 400

            # 百分比归一化
            try:
                payload["seller_percent"] = normalize_pct(payload.get("seller_percent"))
                payload["seller_penalty_ratio"] = normalize_pct(payload.get("seller_penalty_ratio"))
            except ValueError as e:
                return jsonify({"error": str(e)}), 400

            s = Seller(
                seller_code=payload["seller_code"],
                seller_name=payload["seller_name"],
                seller_percent=payload.get("seller_percent"),
                is_catalog_fee_required=bool(payload.get("is_catalog_fee_required", False)),
                is_tax_deductible=bool(payload.get("is_tax_deductible", False)),
                seller_tax_code=payload.get("seller_tax_code"),
                seller_payment_account=payload.get("seller_payment_account"),
                seller_penalty_ratio=payload.get("seller_penalty_ratio"),
                seller_address=payload.get("seller_address"),
                seller_phone=payload.get("seller_phone"),
                seller_notes=payload.get("seller_notes"),
            )
            session.add(s)
            log_op(session, "seller", s.seller_code, "create", before=None, after=str(payload))
            session.commit()
            return jsonify({"ok": True})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # [HTML] 出品人编辑页
    @app.route("/sellers/<seller_code>/edit")
    def sellers_edit(seller_code):
        return render_template("sellers/edit.html", seller_code=seller_code)

    # [API] 更新出品人
    @app.route("/api/sellers/<seller_code>", methods=["PUT"])
    def api_sellers_update(seller_code):
        payload = request.json or {}
        session = get_session()
        try:
            s = session.get(Seller, seller_code)
            if not s:
                return jsonify({"error": "seller 不存在"}), 404
            before = {
                "seller_name": s.seller_name,
                "seller_percent": str(s.seller_percent) if s.seller_percent is not None else None,
                "is_catalog_fee_required": s.is_catalog_fee_required,
                "is_tax_deductible": s.is_tax_deductible,
                "seller_tax_code": s.seller_tax_code,
                "seller_payment_account": s.seller_payment_account,
                "seller_penalty_ratio": str(s.seller_penalty_ratio) if s.seller_penalty_ratio is not None else None,
                "seller_address": s.seller_address,
                "seller_phone": s.seller_phone,
                "seller_notes": s.seller_notes,
            }
            # 若修改名称，需要确保不重名
            new_name = payload.get("seller_name")
            if new_name and new_name != s.seller_name:
                exists_name = session.query(Seller).filter(Seller.seller_name == new_name).first()
                if exists_name:
                    return jsonify({"error": "seller_name 已存在"}), 400
                s.seller_name = new_name

            # 百分比归一化（仅当传入时处理）
            try:
                if "seller_percent" in payload:
                    payload["seller_percent"] = normalize_pct(payload.get("seller_percent"))
                if "seller_penalty_ratio" in payload:
                    payload["seller_penalty_ratio"] = normalize_pct(payload.get("seller_penalty_ratio"))
            except ValueError as e:
                return jsonify({"error": str(e)}), 400

            for k in ["seller_percent", "is_catalog_fee_required", "is_tax_deductible",
                      "seller_tax_code", "seller_payment_account", "seller_penalty_ratio",
                      "seller_address", "seller_phone", "seller_notes"]:
                if k in payload:
                    setattr(s, k, payload[k])

            log_op(session, "seller", s.seller_code, "update", before=str(before), after=str(payload))
            session.commit()
            return jsonify({"ok": True})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    # [API] 删除出品人
    @app.route("/api/sellers/<seller_code>", methods=["DELETE"])
    def api_sellers_delete(seller_code):
        session = get_session()
        try:
            s = session.get(Seller, seller_code)
            if not s:
                return jsonify({"error": "seller 不存在"}), 404
            session.delete(s)
            log_op(session, "seller", seller_code, "delete", before=None, after=None)
            session.commit()
            return jsonify({"ok": True})
        except Exception as e:
            session.rollback()
            return jsonify({"error": str(e)}), 500
        finally:
            session.close()

    return app


# =============================== 入口（开发/部署） ===============================
if __name__ == "__main__":
    app = create_app()
    app.run(host=HOST, port=PORT, debug=DEBUG)
